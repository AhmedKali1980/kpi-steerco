import argparse
import base64
import gzip
import json
import logging
import os
import re
import subprocess
import sys
import unicodedata
import warnings
import zipfile
import shutil
import xml.etree.ElementTree as ET
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.email_utils import parse_recipients, send_carto_notification

try:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError:  # pragma: no cover - runtime dependency check
    load_workbook = None
    BarChart = None
    Reference = None
    DataLabelList = None

try:  # pragma: no cover - optional runtime dependency for PNG charts
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except ImportError:  # pragma: no cover - optional runtime dependency for PNG charts
    plt = None


KPI_MAIL_SHEETS = [
    ("Summary", "summary"),
    ("SCOPE",),
    ("STATS",),
    ("TOTAL.PROGRAM",),
    ("TOTAL.ENTITY",),
    ("NOT_IN_ILLUMIO",),
    ("IN_ILLUMIO_BUT_NOT_BLOCKING",),
    ("EXCLUDED",),
    ("GLOBAL",),
    ("OUT_OF_SCOPE",),
    ("MONITORED_SCOPES",),
]

SMTP_CONF_KEYS = [
    "SMTP_SERVER",
    "SMTP_PORT",
    "SMTP_USE_TLS",
    "SMTP_USE_SSL",
    "SMTP_TIMEOUT",
    "SMTP_USER",
    "SMTP_PASSWORD",
    "SMTP_FROM",
    "SMTP_REPLY_TO",
]

PROGRAM_CHARTS_SHEET = "PROGRAM_CHARTS"


def load_env_file(env_file: str = ".env") -> None:
    path = Path(env_file)
    if not path.is_file():
        return
    with open(path, "r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def setup_logging(log_file: Path, verbose: bool) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    level = logging.DEBUG if verbose else logging.INFO
    root = logging.getLogger()
    root.handlers.clear()
    root.setLevel(level)

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(name)s | %(message)s")

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(level)
    file_handler.setFormatter(formatter)
    root.addHandler(file_handler)

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setLevel(level)
    stream_handler.setFormatter(formatter)
    root.addHandler(stream_handler)


def ensure_inputs_exist(paths: List[Path]) -> None:
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise FileNotFoundError(f"Missing required input files: {', '.join(missing)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="KPI project orchestrator (initial DALI extraction step).")
    parser.add_argument("--runs-dir", default="RUNS", help="Base directory for timestamped runs")
    parser.add_argument("--monitored-file", default="user_inputs/monitored_kears.csv", help="Monitored KEAR input CSV")
    parser.add_argument("--headers-file", default="user_inputs/headers.csv", help="Headers mapping CSV")
    parser.add_argument("--filters-file", default="user_inputs/filters.conf", help="Custom filters conf file")
    parser.add_argument("--dry-run", action="store_true", help="Run extraction without calling DALI API")
    parser.add_argument(
        "--pce-stub-dir",
        default="",
        help="Use existing PCE CSV files from this directory instead of running live exports",
    )
    parser.add_argument("--skip-pce-import", action="store_true", help="Skip PCE workload/iplist import step")
    parser.add_argument("--verbose", action="store_true", help="Verbose logs")
    return parser.parse_args()


def run_pce_import(run_dir: Path, raw_dir: Path, stub_dir: str, log: logging.Logger) -> None:
    cmd = ["bash", "bin/cron_job.sh", str(run_dir)]
    env = os.environ.copy()
    if stub_dir:
        env["PCE_STUB_DIR"] = stub_dir

    log.info("Prepared PCE import command: %s", " ".join(cmd))
    if stub_dir:
        log.info("PCE import mode: stub (source=%s)", stub_dir)
    else:
        log.info("PCE import mode: live")

    result = subprocess.run(cmd, capture_output=True, text=True, env=env)
    if result.stdout:
        log.info("pce import stdout:\n%s", result.stdout.strip())
    if result.stderr:
        log.warning("pce import stderr:\n%s", result.stderr.strip())
    if result.returncode != 0:
        log.error("PCE import failed with exit code %s", result.returncode)
        raise SystemExit(result.returncode)

    expected_files = [raw_dir / "export_wkld.csv", raw_dir / "export_iplists.csv"]
    missing = [str(path) for path in expected_files if not path.is_file() or path.stat().st_size == 0]
    if missing:
        log.error("PCE import completed but missing/empty files: %s", ", ".join(missing))
        raise SystemExit(2)

    log.info("PCE import completed: %s", ", ".join(str(path) for path in expected_files))


def build_mail_conf_from_env() -> Dict[str, str]:
    return {key: (os.getenv(key) or "").strip() for key in SMTP_CONF_KEYS}


def find_generated_pptx(output_xlsx: Path, raw_dir: Path, log: logging.Logger) -> Path:
    preferred = output_xlsx.with_suffix(".pptx")
    if preferred.is_file():
        return preferred

    candidates = sorted(
        raw_dir.glob("*.pptx"),
        key=lambda path: (path.stat().st_mtime, path.name),
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"Missing generated PPTX output in {raw_dir}")

    if len(candidates) > 1:
        log.warning(
            "Multiple PPTX files found in %s; using newest candidate: %s",
            raw_dir,
            candidates[0],
        )
    return candidates[0]


def rename_generated_pptx_for_delivery(
    output_xlsx: Path,
    raw_dir: Path,
    timestamp: str,
    log: logging.Logger,
) -> Path:
    source_pptx = find_generated_pptx(output_xlsx=output_xlsx, raw_dir=raw_dir, log=log)
    target_pptx = raw_dir / f"kpi_microseg_slides_{timestamp}.pptx"

    if source_pptx.resolve() == target_pptx.resolve():
        return source_pptx

    if target_pptx.exists():
        target_pptx.unlink()

    try:
        source_pptx.replace(target_pptx)
        log.info("Renamed PPTX output for delivery: %s -> %s", source_pptx.name, target_pptx.name)
        return target_pptx
    except Exception:
        log.exception("Unable to rename PPTX output for delivery; keeping original PPTX name")
        return source_pptx


def resolve_requested_sheet_names(
    source_sheet_names: List[str],
    requested_sheets: List[Tuple[str, ...]],
) -> List[str]:
    resolved: List[str] = []
    source_lower_map = {sheet_name.lower(): sheet_name for sheet_name in source_sheet_names}

    for candidates in requested_sheets:
        matched_sheet = None

        for candidate in candidates:
            if candidate in source_sheet_names:
                matched_sheet = candidate
                break

        if matched_sheet is None:
            for candidate in candidates:
                matched_sheet = source_lower_map.get(candidate.lower())
                if matched_sheet is not None:
                    break

        if matched_sheet is None:
            raise ValueError(
                "Unable to prepare KPI email workbook. Missing sheets in source XLSX: "
                + " / ".join(candidates)
            )

        resolved.append(matched_sheet)

    return resolved


def _load_workbook_with_warning_filter(path: Path, *, data_only: bool = False):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to prepare the email Excel attachment")

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Conditional Formatting extension is not supported and will be removed",
            category=UserWarning,
        )
        return load_workbook(filename=path, data_only=data_only)


def _xlsx_sheet_prune_copy(
    source_xlsx: Path,
    destination_xlsx: Path,
    keep_sheet_names: List[Tuple[str, ...]],
    log: logging.Logger,
) -> List[str]:
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg_rel = "http://schemas.openxmlformats.org/package/2006/relationships"
    ns_ct = "http://schemas.openxmlformats.org/package/2006/content-types"
    ns_vt = "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"

    ET.register_namespace("", ns_main)
    ET.register_namespace("r", ns_rel)

    destination_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_xlsx, destination_xlsx)

    with zipfile.ZipFile(destination_xlsx, "r") as zin:
        original_entries = {info.filename: zin.read(info.filename) for info in zin.infolist()}

    workbook_xml = ET.fromstring(original_entries["xl/workbook.xml"])
    workbook_rels_xml = ET.fromstring(original_entries["xl/_rels/workbook.xml.rels"])
    app_xml = ET.fromstring(original_entries.get("docProps/app.xml", b"<Properties/>"))
    ct_xml = ET.fromstring(original_entries["[Content_Types].xml"])

    sheets_parent = workbook_xml.find(f"{{{ns_main}}}sheets")
    if sheets_parent is None:
        raise ValueError("xl/workbook.xml does not contain a sheets collection")

    source_sheets = [sheet.attrib.get("name", "") for sheet in sheets_parent.findall(f"{{{ns_main}}}sheet")]
    resolved_sheet_names = resolve_requested_sheet_names(source_sheets, keep_sheet_names)
    keep_set = set(resolved_sheet_names)

    rels_by_id = {rel.attrib.get("Id"): rel for rel in workbook_rels_xml.findall(f"{{{ns_pkg_rel}}}Relationship")}
    removed_sheet_paths = set()

    original_sheet_elements = list(sheets_parent.findall(f"{{{ns_main}}}sheet"))
    for sheet in original_sheet_elements:
        name = sheet.attrib.get("name", "")
        rid = sheet.attrib.get(f"{{{ns_rel}}}id")
        rel = rels_by_id.get(rid)
        if name not in keep_set:
            sheets_parent.remove(sheet)
            if rel is not None:
                target = rel.attrib.get("Target", "")
                if target:
                    target_path = "xl/" + target.lstrip("/") if not target.startswith("/") else target.lstrip("/")
                    removed_sheet_paths.add(target_path)
                    removed_sheet_paths.add(str(Path(target_path).parent / "_rels" / (Path(target_path).name + ".rels")))
                workbook_rels_xml.remove(rel)

    kept_sheet_elements = list(sheets_parent.findall(f"{{{ns_main}}}sheet"))
    old_index_by_name = {sheet.attrib.get("name", ""): idx for idx, sheet in enumerate(original_sheet_elements)}
    new_index_by_name = {sheet.attrib.get("name", ""): idx for idx, sheet in enumerate(kept_sheet_elements)}

    for idx, sheet in enumerate(kept_sheet_elements, start=1):
        sheet.attrib["sheetId"] = str(idx)

    book_views = workbook_xml.find(f"{{{ns_main}}}bookViews")
    if book_views is not None:
        workbook_view = book_views.find(f"{{{ns_main}}}workbookView")
        if workbook_view is not None:
            active_tab = workbook_view.attrib.get("activeTab")
            if active_tab is not None:
                try:
                    active_tab_int = int(active_tab)
                except Exception:
                    active_tab_int = 0
                active_tab_int = max(0, min(active_tab_int, max(0, len(kept_sheet_elements) - 1)))
                workbook_view.attrib["activeTab"] = str(active_tab_int)

    defined_names = workbook_xml.find(f"{{{ns_main}}}definedNames")
    if defined_names is not None:
        for defined_name in list(defined_names.findall(f"{{{ns_main}}}definedName")):
            local_sheet_id = defined_name.attrib.get("localSheetId")
            if local_sheet_id is None:
                continue
            try:
                old_local_idx = int(local_sheet_id)
            except Exception:
                continue
            matched_name = None
            for name, old_idx in old_index_by_name.items():
                if old_idx == old_local_idx:
                    matched_name = name
                    break
            if matched_name not in new_index_by_name:
                defined_names.remove(defined_name)
            else:
                defined_name.attrib["localSheetId"] = str(new_index_by_name[matched_name])

    titles_of_parts = app_xml.find(f".//{{{ns_vt}}}vector")
    if titles_of_parts is not None:
        lpstr_nodes = titles_of_parts.findall(f"{{{ns_vt}}}lpstr")
        if lpstr_nodes:
            for node in list(lpstr_nodes):
                titles_of_parts.remove(node)
            for name in resolved_sheet_names:
                elem = ET.Element(f"{{{ns_vt}}}lpstr")
                elem.text = name
                titles_of_parts.append(elem)
            titles_of_parts.attrib["size"] = str(len(resolved_sheet_names))

    heading_pairs = app_xml.find(f".//{{{ns_vt}}}vector")
    if heading_pairs is not None:
        i4_nodes = heading_pairs.findall(f"{{{ns_vt}}}variant/{{{ns_vt}}}i4")
        if len(i4_nodes) >= 2:
            i4_nodes[1].text = str(len(resolved_sheet_names))

    overrides_parent = ct_xml
    for override in list(overrides_parent.findall(f"{{{ns_ct}}}Override")):
        part_name = override.attrib.get("PartName", "").lstrip("/")
        if part_name in removed_sheet_paths:
            overrides_parent.remove(override)

    entries_to_remove = {entry for entry in removed_sheet_paths if entry in original_entries}

    updated_entries = dict(original_entries)
    updated_entries["xl/workbook.xml"] = ET.tostring(workbook_xml, encoding="utf-8", xml_declaration=True)
    updated_entries["xl/_rels/workbook.xml.rels"] = ET.tostring(workbook_rels_xml, encoding="utf-8", xml_declaration=True)
    if "docProps/app.xml" in updated_entries:
        updated_entries["docProps/app.xml"] = ET.tostring(app_xml, encoding="utf-8", xml_declaration=True)
    updated_entries["[Content_Types].xml"] = ET.tostring(ct_xml, encoding="utf-8", xml_declaration=True)

    for entry in entries_to_remove:
        updated_entries.pop(entry, None)

    temp_path = destination_xlsx.with_suffix(destination_xlsx.suffix + ".tmp")
    with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in updated_entries.items():
            zout.writestr(name, data)

    temp_path.replace(destination_xlsx)
    return resolved_sheet_names


def create_email_attachment_workbook(
    source_xlsx: Path,
    destination_xlsx: Path,
    keep_sheet_names: List[Tuple[str, ...]],
    log: logging.Logger,
) -> Path:
    resolved_sheet_names = _xlsx_sheet_prune_copy(
        source_xlsx=source_xlsx,
        destination_xlsx=destination_xlsx,
        keep_sheet_names=keep_sheet_names,
        log=log,
    )

    log.info(
        "Prepared reduced KPI Excel attachment without workbook round-trip: %s (sheets=%s)",
        destination_xlsx,
        ", ".join(resolved_sheet_names),
    )
    return destination_xlsx


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("%", " percent ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def score_program_header(header: str) -> int:
    score = 0
    if "program" in header:
        score += 10
    if header in {"program", "program name"}:
        score += 5
    return score


def score_install_header(header: str) -> int:
    score = 0
    if "installed" in header:
        score += 8
    if "illumio" in header:
        score += 4
    if "percent" in header:
        score += 2
    if "server" in header:
        score += 1
    if "blocking" in header:
        score -= 6
    if "enforced" in header:
        score -= 4
    return score


def score_blocking_header(header: str) -> int:
    score = 0
    if "blocking" in header:
        score += 8
    if "enforced" in header:
        score += 7
    if "agent" in header:
        score += 2
    if "illumio" in header:
        score += 3
    if "percent" in header:
        score += 2
    if "server" in header:
        score += 1
    if "installed" in header and "blocking" not in header:
        score -= 6
    return score


def score_total_servers_header(header: str) -> int:
    score = 0
    if "server" in header:
        score += 4
    if "total" in header:
        score += 8
    if header in {"servers", "server count", "total servers", "nb servers", "number of servers"}:
        score += 6
    if "illumio" in header:
        score -= 6
    if "install" in header:
        score -= 5
    if "blocking" in header or "enforced" in header:
        score -= 5
    if "percent" in header:
        score -= 8
    return score


def score_installed_count_header(header: str) -> int:
    score = 0
    if "install" in header:
        score += 8
    if "illumio" in header:
        score += 5
    if "server" in header:
        score += 2
    if "count" in header or "nb" in header or "number" in header:
        score += 2
    if "percent" in header:
        score -= 8
    if "blocking" in header or "enforced" in header:
        score -= 6
    return score


def score_blocking_count_header(header: str) -> int:
    score = 0
    if "blocking" in header:
        score += 8
    if "enforced" in header:
        score += 7
    if "illumio" in header:
        score += 4
    if "agent" in header:
        score += 2
    if "server" in header:
        score += 2
    if "count" in header or "nb" in header or "number" in header:
        score += 2
    if "percent" in header:
        score -= 8
    if "install" in header and "blocking" not in header and "enforced" not in header:
        score -= 6
    return score


def find_total_program_header_row(ws) -> int:
    best_row = 1
    best_score = -1
    max_scan_row = min(ws.max_row, 10)
    for row_idx in range(1, max_scan_row + 1):
        headers = [normalize_header(cell.value) for cell in ws[row_idx]]
        row_score = max((score_program_header(h) for h in headers), default=0)
        row_score += max((score_install_header(h) for h in headers), default=0)
        row_score += max((score_blocking_header(h) for h in headers), default=0)
        if row_score > best_score:
            best_score = row_score
            best_row = row_idx
    if best_score <= 0:
        raise ValueError("Unable to detect the header row in TOTAL.PROGRAM")
    return best_row


def pick_best_column(headers: dict[int, str], scorer, label: str) -> int:
    best_col = 0
    best_score = -10**9
    for col_idx, header in headers.items():
        score = scorer(header)
        if score > best_score:
            best_score = score
            best_col = col_idx
    if best_col <= 0 or best_score <= 0:
        raise ValueError(f"Unable to detect the '{label}' column in TOTAL.PROGRAM")
    return best_col


def pick_optional_best_column(headers: Dict[int, str], scorer) -> Optional[int]:
    best_col = 0
    best_score = -10**9
    for col_idx, header in headers.items():
        score = scorer(header)
        if score > best_score:
            best_score = score
            best_col = col_idx
    if best_col <= 0 or best_score <= 0:
        return None
    return best_col


def normalize_percentage_value(value: Any) -> Optional[float]:
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None

        percent_matches = re.findall(r"([-+]?\d+(?:[.,]\d+)?)\s*%", stripped)
        if percent_matches:
            try:
                return round(float(percent_matches[-1].replace(",", ".")), 2)
            except ValueError:
                pass

        ratio_match = re.search(r"(\d+(?:[.,]\d+)?)\s*/\s*(\d+(?:[.,]\d+)?)", stripped)
        if ratio_match:
            try:
                numerator = float(ratio_match.group(1).replace(",", "."))
                denominator = float(ratio_match.group(2).replace(",", "."))
                if denominator != 0:
                    return round((numerator / denominator) * 100.0, 2)
            except ValueError:
                pass

        cleaned = stripped.replace("(", "").replace(")", "").replace("%", "").replace(",", ".").strip()
        if not cleaned:
            return None
        try:
            numeric = float(cleaned)
        except ValueError:
            return None
    else:
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return None

    if abs(numeric) <= 1.0:
        numeric *= 100.0
    return round(numeric, 2)


def normalize_numeric_value(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        stripped = stripped.replace(",", ".")
        try:
            numeric = float(stripped)
        except ValueError:
            return None
    else:
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return None
    return float(numeric)


def extract_total_program_metrics(source_xlsx: Path, log: logging.Logger) -> List[Dict[str, Union[float, str]]]:
    workbook = _load_workbook_with_warning_filter(source_xlsx, data_only=False)
    try:
        total_program_sheet_name = resolve_requested_sheet_names(list(workbook.sheetnames), [("TOTAL.PROGRAM",)])[0]
        ws = workbook[total_program_sheet_name]
        header_row = find_total_program_header_row(ws)
        headers = {col_idx: normalize_header(ws.cell(header_row, col_idx).value) for col_idx in range(1, ws.max_column + 1)}
        program_col = pick_best_column(headers, score_program_header, "program")
        install_col = pick_best_column(headers, score_install_header, "installed percentage")
        blocking_col = pick_best_column(headers, score_blocking_header, "blocking percentage")
        total_servers_col = pick_optional_best_column(headers, score_total_servers_header)
        installed_count_col = pick_optional_best_column(headers, score_installed_count_header)
        blocking_count_col = pick_optional_best_column(headers, score_blocking_count_header)

        log.info(
            "TOTAL.PROGRAM header row=%s program_col=%s install_pct_col=%s blocking_pct_col=%s total_servers_col=%s installed_count_col=%s blocking_count_col=%s",
            header_row,
            program_col,
            install_col,
            blocking_col,
            total_servers_col,
            installed_count_col,
            blocking_count_col,
        )

        rows: List[Dict[str, Union[float, str]]] = []
        direct_value_rows = 0
        computed_rows = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            program_value = ws.cell(row_idx, program_col).value
            program_name = str(program_value or "").strip()
            if not program_name:
                continue
            if program_name.lower() in {"total", "totals", "grand total"}:
                continue

            install_pct = normalize_percentage_value(ws.cell(row_idx, install_col).value)
            blocking_pct = normalize_percentage_value(ws.cell(row_idx, blocking_col).value)

            total_servers = None
            installed_count = None
            blocking_count = None
            if total_servers_col is not None:
                total_servers = normalize_numeric_value(ws.cell(row_idx, total_servers_col).value)
            if installed_count_col is not None:
                installed_count = normalize_numeric_value(ws.cell(row_idx, installed_count_col).value)
            if blocking_count_col is not None:
                blocking_count = normalize_numeric_value(ws.cell(row_idx, blocking_count_col).value)

            if install_pct is None and total_servers not in (None, 0) and installed_count is not None:
                install_pct = round((installed_count / total_servers) * 100.0, 2)
            if blocking_pct is None and total_servers not in (None, 0) and blocking_count is not None:
                blocking_pct = round((blocking_count / total_servers) * 100.0, 2)

            if install_pct is None and blocking_pct is None:
                continue

            if normalize_percentage_value(ws.cell(row_idx, install_col).value) is not None or normalize_percentage_value(ws.cell(row_idx, blocking_col).value) is not None:
                direct_value_rows += 1
            else:
                computed_rows += 1

            rows.append(
                {
                    "program": program_name,
                    "installed_pct": install_pct if install_pct is not None else 0.0,
                    "blocking_pct": blocking_pct if blocking_pct is not None else 0.0,
                }
            )
    finally:
        workbook.close()

    if not rows:
        header_debug = ", ".join(f"{col}:{name}" for col, name in headers.items()) if 'headers' in locals() else "<unavailable>"
        raise ValueError(
            "No usable TOTAL.PROGRAM rows were found to build KPI charts. "
            f"Detected headers: {header_debug}"
        )

    log.info(
        "Extracted %s program KPI rows from TOTAL.PROGRAM (direct_value_rows=%s computed_rows=%s)",
        len(rows),
        direct_value_rows,
        computed_rows,
    )
    return rows


def style_program_charts_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws["A1"] = "KPI charts by program"
    ws["A2"] = "Source: TOTAL.PROGRAM"
    ws["A4"] = "Program"
    ws["B4"] = "Installed current (%)"
    ws["C4"] = "Installed target (%)"
    ws["D4"] = "Blocking current (%)"
    ws["E4"] = "Blocking target (%)"

    title_font = copy(ws["A1"].font)
    title_font.bold = True
    title_font.size = 14
    title_font.color = "FFFFFF"
    ws["A1"].font = title_font
    title_fill = copy(ws["A1"].fill)
    title_fill.fill_type = "solid"
    title_fill.fgColor = "1F4E78"
    ws["A1"].fill = title_fill

    subtitle_font = copy(ws["A2"].font)
    subtitle_font.italic = True
    subtitle_font.color = "666666"
    ws["A2"].font = subtitle_font

    for cell in ws[4]:
        header_font = copy(cell.font)
        header_font.bold = True
        header_font.color = "FFFFFF"
        cell.font = header_font
        header_fill = copy(cell.fill)
        header_fill.fill_type = "solid"
        header_fill.fgColor = "1F1F1F"
        cell.fill = header_fill

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 19
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20


def build_openpyxl_column_chart(
    ws,
    title: str,
    current_col: int,
    target_col: int,
    max_row: int,
    anchor: str,
    current_color: str,
    target_color: str,
) -> None:
    if BarChart is None or Reference is None:
        raise RuntimeError("openpyxl chart support is unavailable")

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = title
    chart.y_axis.title = "Percentage"
    chart.x_axis.title = "Program"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.height = 9.5
    chart.width = max(14.0, min(24.0, 1.7 * max(max_row - 4, 1)))
    chart.legend.position = "r"
    chart.varyColors = False

    current_data = Reference(ws, min_col=current_col, min_row=4, max_row=max_row)
    target_data = Reference(ws, min_col=target_col, min_row=4, max_row=max_row)
    categories = Reference(ws, min_col=1, min_row=5, max_row=max_row)
    chart.add_data(target_data, titles_from_data=True)
    chart.add_data(current_data, titles_from_data=True)
    chart.set_categories(categories)

    try:
        chart.gapWidth = 60
    except Exception:
        pass

    if len(chart.series) >= 2:
        target_series = chart.series[0]
        current_series = chart.series[1]
        for series, color in ((target_series, target_color), (current_series, current_color)):
            try:
                series.graphicalProperties.solidFill = color
                series.graphicalProperties.line.solidFill = color
            except Exception:
                pass
        try:
            labels = DataLabelList()
            labels.showVal = True
            labels.position = "outEnd"
            current_series.dLbls = labels
        except Exception:
            pass

    ws.add_chart(chart, anchor)


def add_program_charts_to_workbook(workbook_path: Path, chart_rows: List[Dict[str, Union[float, str]]], log: logging.Logger) -> None:
    workbook = _load_workbook_with_warning_filter(workbook_path)
    try:
        if PROGRAM_CHARTS_SHEET in workbook.sheetnames:
            del workbook[PROGRAM_CHARTS_SHEET]
        ws = workbook.create_sheet(PROGRAM_CHARTS_SHEET)
        style_program_charts_sheet(ws)

        sorted_rows = sorted(chart_rows, key=lambda entry: str(entry["program"]).lower())
        current_row = 5
        for entry in sorted_rows:
            ws.cell(current_row, 1, str(entry["program"]))
            ws.cell(current_row, 2, float(entry["installed_pct"]))
            ws.cell(current_row, 3, 100.0)
            ws.cell(current_row, 4, float(entry["blocking_pct"]))
            ws.cell(current_row, 5, 100.0)
            for col_idx in (2, 3, 4, 5):
                ws.cell(current_row, col_idx).number_format = '0.0"%"'
            current_row += 1

        data_end_row = current_row - 1
        if data_end_row < 5:
            raise ValueError("PROGRAM_CHARTS cannot be built because no chart data rows were prepared")

        build_openpyxl_column_chart(
            ws=ws,
            title="Installed coverage by program (target vs current)",
            current_col=2,
            target_col=3,
            max_row=data_end_row,
            anchor="G4",
            current_color="4F81BD",
            target_color="C9C9C9",
        )
        second_chart_anchor_row = max(24, 12 + len(sorted_rows))
        build_openpyxl_column_chart(
            ws=ws,
            title="Blocking coverage by program (target vs current)",
            current_col=4,
            target_col=5,
            max_row=data_end_row,
            anchor=f"G{second_chart_anchor_row}",
            current_color="9BBB59",
            target_color="C9C9C9",
        )

        workbook.save(workbook_path)
    finally:
        workbook.close()

    log.info("Added %s sheet to %s", PROGRAM_CHARTS_SHEET, workbook_path)


def _build_chart_data_uri(path: Path) -> Optional[str]:
    if not path.is_file():
        return None
    mime = "image/png"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def create_bar_chart_png(
    entries: List[Dict[str, Union[float, str]]],
    metric_key: str,
    title: str,
    output_path: Path,
    color: str,
    target_color: str = "#D9D9D9",
) -> Optional[Path]:
    if plt is None:
        return None

    ordered_entries = sorted(entries, key=lambda entry: float(entry[metric_key]), reverse=True)
    programs = [str(entry["program"]) for entry in ordered_entries]
    values = [float(entry[metric_key]) for entry in ordered_entries]
    targets = [100.0 for _ in ordered_entries]
    positions = list(range(len(programs)))
    width = 0.38

    figure_width = max(10.0, min(24.0, 1.5 * max(len(programs), 1) + 4.0))
    fig, ax = plt.subplots(figsize=(figure_width, 6.8))
    try:
        ax.bar([p - width / 2 for p in positions], targets, width=width, color=target_color, label="Reference target (100%)")
        bars = ax.bar([p + width / 2 for p in positions], values, width=width, color=color, label="Current value")
        ax.set_title(title, fontsize=14, fontweight="bold")
        ax.set_ylabel("Percentage")
        ax.set_xlabel("Program")
        ax.set_ylim(0, 110)
        ax.set_xticks(positions)
        ax.set_xticklabels(programs, rotation=25, ha="right")
        ax.grid(axis="y", linestyle="--", alpha=0.35)
        ax.set_axisbelow(True)
        ax.legend()

        for bar, value in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, min(value + 2.0, 108.0), f"{value:.1f}%", ha="center", va="bottom", fontsize=9)

        fig.tight_layout()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(output_path, dpi=180, bbox_inches="tight")
        return output_path
    finally:
        plt.close(fig)


def prepare_total_program_chart_assets(
    slim_xlsx_path: Path,
    chart_rows: List[Dict[str, Union[float, str]]],
    raw_dir: Path,
    timestamp: str,
    log: logging.Logger,
) -> List[Dict[str, str]]:
    add_program_charts_to_workbook(slim_xlsx_path, chart_rows, log)

    inline_images: List[Dict[str, str]] = []
    image_specs = [
        (
            "installed_by_program",
            "installed_pct",
            "% servers with illumio installed by program",
            raw_dir / f"kpi_microseg_installed_{timestamp}.png",
            "#4F81BD",
        ),
        (
            "blocking_by_program",
            "blocking_pct",
            "% servers with illumio agent in blocking mode by program",
            raw_dir / f"kpi_microseg_blocking_{timestamp}.png",
            "#9BBB59",
        ),
    ]

    for content_id, metric_key, title, output_path, color in image_specs:
        image_path = create_bar_chart_png(
            entries=chart_rows,
            metric_key=metric_key,
            title=title,
            output_path=output_path,
            color=color,
        )
        if image_path is not None and image_path.is_file():
            inline_images.append({"cid": content_id, "path": str(image_path), "title": title, "data_uri": _build_chart_data_uri(image_path) or ""})

    if inline_images:
        log.info(
            "Prepared inline KPI program charts: %s",
            ", ".join(Path(item["path"]).name for item in inline_images),
        )
    else:
        log.warning("Program chart PNG generation skipped because matplotlib is unavailable")

    return inline_images


def build_kpi_mail_bodies(
    timestamp: str,
    meta: Dict[str, Any],
    pptx_path: Path,
    slim_xlsx_path: Path,
    inline_images: Optional[List[Dict[str, str]]] = None,
) -> Tuple[str, str]:
    uid_count = int(meta.get("uid_count", 0) or 0)
    success_count = int(meta.get("success_count", 0) or 0)
    found_count = int(meta.get("found_count", 0) or 0)
    error_count = int(meta.get("error_count", 0) or 0)

    body_text_lines = [
        "Hello,",
        "",
        f"Please find attached the KPI microsegmentation reports generated for run {timestamp}.",
        "",
        "Attachments:",
        f"- Full PowerPoint report: {pptx_path.name}",
        f"- Reduced Excel report: {slim_xlsx_path.name}",
        "",
        "DALI summary:",
        f"- uid_count: {uid_count}",
        f"- success_count: {success_count}",
        f"- found_count: {found_count}",
        f"- error_count: {error_count}",
        "",
        "Regards,",
        "kpi_orchestrator",
    ]
    body_text = "\n".join(body_text_lines)

    body_html_parts = [
        "<p>Hello,</p>",
        f"<p>Please find attached the KPI microsegmentation reports generated for run <strong>{timestamp}</strong>.</p>",
        "<p>Attachments:</p>",
        "<ul>",
        f"  <li>Full PowerPoint report: <strong>{pptx_path.name}</strong></li>",
        f"  <li>Reduced Excel report: <strong>{slim_xlsx_path.name}</strong></li>",
        "</ul>",
        "<p>DALI summary:</p>",
        "<ul>",
        f"  <li>uid_count: {uid_count}</li>",
        f"  <li>success_count: {success_count}</li>",
        f"  <li>found_count: {found_count}</li>",
        f"  <li>error_count: {error_count}</li>",
        "</ul>",
        "<p>Regards,<br>kpi_orchestrator</p>",
    ]
    body_html = "\n".join(body_html_parts)
    return body_text, body_html


def maybe_send_kpi_email(
    timestamp: str,
    raw_dir: Path,
    output_xlsx: Path,
    meta: Dict[str, Any],
    log: logging.Logger,
) -> None:
    recipients = parse_recipients(os.getenv("MAIL_TO", ""))
    if not recipients:
        log.info("MAIL_TO is empty; KPI email notification skipped.")
        return

    mail_conf = build_mail_conf_from_env()
    if not mail_conf["SMTP_SERVER"]:
        log.warning("MAIL_TO is set but SMTP_SERVER is empty; KPI email notification skipped.")
        return

    pptx_path = rename_generated_pptx_for_delivery(
        output_xlsx=output_xlsx,
        raw_dir=raw_dir,
        timestamp=timestamp,
        log=log,
    )

    slim_xlsx_path = raw_dir / f"kpi_microseg_{timestamp}.xlsx"
    create_email_attachment_workbook(
        source_xlsx=output_xlsx,
        destination_xlsx=slim_xlsx_path,
        keep_sheet_names=KPI_MAIL_SHEETS,
        log=log,
    )

    subject = f"KPI Microseg report - {timestamp}"
    body_text, body_html = build_kpi_mail_bodies(
        timestamp=timestamp,
        meta=meta,
        pptx_path=pptx_path,
        slim_xlsx_path=slim_xlsx_path,
        inline_images=None,
    )

    send_carto_notification(
        conf=mail_conf,
        recipients=recipients,
        subject=subject,
        body_text=body_text,
        body_html=body_html,
        attachment_paths=[pptx_path, slim_xlsx_path],
        inline_images=None,
        logger=log,
    )

    log.info(
        "KPI notification email sent with attachments: %s, %s",
        pptx_path.name,
        slim_xlsx_path.name,
    )


def main() -> None:
    args = parse_args()
    load_env_file()

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    run_dir = Path(args.runs_dir) / timestamp
    raw_dir = run_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    log_file = run_dir / "execution.log"
    setup_logging(log_file=log_file, verbose=args.verbose)
    log = logging.getLogger("kpi_orchestrator")

    log.info("Starting KPI orchestration")
    log.info("Run directory initialized: %s", run_dir)
    log.info("Raw directory initialized: %s", raw_dir)

    monitored_file = Path(args.monitored_file)
    headers_file = Path(args.headers_file)
    filters_file = Path(args.filters_file)
    ensure_inputs_exist([monitored_file, headers_file, filters_file])
    log.info("Validated user inputs: %s, %s, %s", monitored_file, headers_file, filters_file)

    if args.skip_pce_import:
        log.info("PCE import skipped by --skip-pce-import")
    else:
        run_pce_import(run_dir=run_dir, raw_dir=raw_dir, stub_dir=args.pce_stub_dir.strip(), log=log)

    internet_exposed_xlsx = raw_dir / f"internet_exposed_{timestamp}.xlsx"
    internet_exposed_csv = raw_dir / f"internet_exposed_{timestamp}.csv"
    internet_exposed_json = raw_dir / "internet_exposed.json"
    internet_cmd = [
        sys.executable,
        "modules/internet_exposed_extract.py",
        "--output",
        str(internet_exposed_xlsx),
        "--csv-out",
        str(internet_exposed_csv),
        "--json-out",
        str(internet_exposed_json),
        "--filters-file",
        str(filters_file),
    ]
    if args.verbose:
        internet_cmd.append("--verbose")
    log.info("Prepared INTERNET.EXPOSED extraction command: %s", " ".join(internet_cmd))
    internet_result = subprocess.run(internet_cmd, capture_output=True, text=True)
    if internet_result.stdout:
        log.info("internet_exposed_extract stdout:\n%s", internet_result.stdout.strip())
    if internet_result.stderr:
        log.warning("internet_exposed_extract stderr:\n%s", internet_result.stderr.strip())
    if internet_result.returncode != 0:
        log.error("internet_exposed_extract.py failed with exit code %s", internet_result.returncode)
        raise SystemExit(internet_result.returncode)
    if not internet_exposed_xlsx.is_file() or not Path(str(internet_exposed_json) + ".gz").is_file():
        log.error("Expected INTERNET.EXPOSED output files missing in %s", raw_dir)
        raise SystemExit(2)
    log.info("INTERNET.EXPOSED XLSX output: %s", internet_exposed_xlsx)

    depth_until = (os.getenv("DALI_DEPTH_UNTIL") or "").strip()
    limit = (os.getenv("DALI_LIMIT") or "").strip()
    impact_endpoint = (os.getenv("DALI_IMPACT_ENDPOINT") or "/api/v1/impactAnalysis").strip()

    if not depth_until:
        log.warning("DALI_DEPTH_UNTIL is not set in .env; default from dali_impact_analysis.py will apply.")
    if not limit:
        log.warning("DALI_LIMIT is not set in .env; default from dali_impact_analysis.py will apply.")
    if not impact_endpoint:
        impact_endpoint = "/api/v1/impactAnalysis"

    if "xxxxxxxx" in (os.getenv("SGCONNECT_CLIENT_ID") or "") or "xxxxxxxx" in (os.getenv("SGCONNECT_CLIENT_SECRET") or ""):
        log.warning("SGCONNECT credentials appear to be placeholders; DALI live calls may fail.")

    output_xlsx = raw_dir / f"dali_impact_analysis_{timestamp}.xlsx"
    output_json = raw_dir / "dali_impact_analysis.json"
    output_json_gz = Path(str(output_json) + ".gz")

    cmd = [
        sys.executable,
        "modules/dali_impact_analysis.py",
        "--monitored-file",
        str(monitored_file),
        "--headers-file",
        str(headers_file),
        "--filters-file",
        str(filters_file),
        "--output",
        str(output_xlsx),
        "--json-out",
        str(output_json),
    ]

    if depth_until:
        cmd.extend(["--depth-until", depth_until])
    if limit:
        cmd.extend(["--limit", limit])
    cmd.extend(["--impact-endpoint", impact_endpoint])
    if args.dry_run:
        cmd.append("--dry-run")
    if args.verbose:
        cmd.append("--verbose")

    log.info("Prepared DALI extraction command: %s", " ".join(cmd))

    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.stdout:
        log.info("dali_impact_analysis stdout:\n%s", result.stdout.strip())
    if result.stderr:
        log.warning("dali_impact_analysis stderr:\n%s", result.stderr.strip())

    if result.returncode != 0:
        log.error("dali_impact_analysis.py failed with exit code %s", result.returncode)
        raise SystemExit(result.returncode)

    if not output_xlsx.is_file() or not output_json_gz.is_file():
        log.error("Expected output files missing in %s", raw_dir)
        raise SystemExit(2)

    payload: Dict[str, Any] = {}
    meta: Dict[str, Any] = {}
    try:
        with gzip.open(output_json_gz, "rt", encoding="utf-8") as handle:
            payload = json.load(handle)
        meta = payload.get("meta", {}) if isinstance(payload, dict) else {}
        uid_count = int(meta.get("uid_count", 0) or 0)
        success_count = int(meta.get("success_count", 0) or 0)
        error_count = int(meta.get("error_count", 0) or 0)
        found_count = int(meta.get("found_count", 0) or 0)

        log.info(
            "DALI summary: uid_count=%s success_count=%s found_count=%s error_count=%s",
            uid_count,
            success_count,
            found_count,
            error_count,
        )

        if not args.dry_run and uid_count > 0 and success_count == 0:
            errors = payload.get("errors", []) if isinstance(payload, dict) else []
            all_http_400 = bool(errors) and all("HTTP 400" in str(err.get("error", "")) for err in errors if isinstance(err, dict))
            for err in errors[:3]:
                if isinstance(err, dict):
                    log.error("DALI error detail uid=%s: %s", err.get("uid", "<unknown>"), err.get("error", ""))
            if all_http_400:
                log.error(
                    "All DALI requests failed with HTTP 400. TLS seems configured; check impact endpoint and query params (filters/status/zones/environments) against DALI API contract."
                )
            else:
                log.error(
                    "All DALI requests failed (0 successful calls). Check TLS/CA config (VERIFY_CA or SG CA bundle), credentials, and API parameters."
                )
            raise SystemExit(3)
    except SystemExit:
        raise
    except Exception as exc:
        log.warning("Unable to parse JSON summary for post-check: %s", exc)

    try:
        pptx_output = rename_generated_pptx_for_delivery(
            output_xlsx=output_xlsx,
            raw_dir=raw_dir,
            timestamp=timestamp,
            log=log,
        )
        log.info("PPTX output: %s", pptx_output)
    except FileNotFoundError as exc:
        log.warning(str(exc))

    maybe_send_kpi_email(
        timestamp=timestamp,
        raw_dir=raw_dir,
        output_xlsx=output_xlsx,
        meta=meta,
        log=log,
    )

    log.info("DALI extraction completed successfully")
    log.info("XLSX output: %s", output_xlsx)
    log.info("JSON.GZ output: %s", output_json_gz)
    log.info("Execution log: %s", log_file)


if __name__ == "__main__":
    main()
