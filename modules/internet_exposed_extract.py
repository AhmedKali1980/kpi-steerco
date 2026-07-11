import argparse
import csv
import gzip
import json
import logging
import os
import random
import re
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple
from urllib.parse import urljoin

from elasticsearch.helpers import scan

from config import QUERY_CONFIG
from d4s_client import Data4secClient
from dali_impact_analysis import DaliImpactAnalysisClient

log = logging.getLogger(__name__)
TECHNICAL_FIELDS = ["exposure_scopes", "is_dali_exposed", "is_masai_exposed"]
ALL_FILTERS_FIELD = "F_ALL_FILTERS"
MARLEY_KEAR_UUID_FIELD = "MAR_app_info.kear_uuid"
MARLEY_KEAR_FACTOR_FIELD = "MAR_app_info.kear_factor"
CALCULATED_SINGLE_KEAR_FIELD = "calculated_Single_Kear"
MISSING_KEAR_VALUE = "MISSING_KEAR"
MARLEY_KEAR_FIELDS = [MARLEY_KEAR_UUID_FIELD, MARLEY_KEAR_FACTOR_FIELD, CALCULATED_SINGLE_KEAR_FIELD]

PCE_WORKLOAD_FIELDS = [
    "PCE_match_status",
    "PCE_match_method",
    "PCE_hostname",
    "PCE_short_hostname",
    "PCE_created_at",
    "PCE_ip_with_default_gw",
    "PCE_app",
    "PCE_env",
    "PCE_role",
    "PCE_loc",
    "PCE_OS",
    "PCE_managed",
    "PCE_enforcement",
    "PCE_ocs_name_from_IP",
    "PCE_IPLIST",
    "PCE_SUBNET",
]
PCE_WORKLOAD_SOURCE_FIELDS = {
    "PCE_hostname": ["hostname", "hosname"],
    "PCE_short_hostname": ["short_hostname"],
    "PCE_created_at": ["created_at"],
    "PCE_ip_with_default_gw": ["ip_with_default_gw"],
    "PCE_app": ["app"],
    "PCE_env": ["env"],
    "PCE_role": ["role"],
    "PCE_loc": ["loc"],
    "PCE_OS": ["OS", "os"],
    "PCE_managed": ["managed"],
    "PCE_enforcement": ["enforcement"],
    "PCE_ocs_name_from_IP": ["ocs_name_from_IP"],
    "PCE_IPLIST": ["IPLIST"],
    "PCE_SUBNET": ["SUBNET"],
}
_IP_DERIVED_NAME_RE = re.compile(r"^(?:IP-)?\d{1,3}(?:-\d{1,3}){3}$", re.IGNORECASE)
CALCULATED_ENV_FILTER_FIELD = "F_env_calculated"
INVENTORY_ENRICHMENT_FIELDS = [
    "INV_owner_app_name",
    "PA_owner_id",
    "INV_beneficiary",
    "PA_beneficiary_id",
    "PA_beneficiary_ENV",
    "INV_region",
    CALCULATED_ENV_FILTER_FIELD,
]
DICT_ACCOUNT_HEADERS = ["account", "id", "env"]
DICT_DALI_APP_SHEET = "DictDaliApp"
DICT_DALI_APP_SOURCE_FIELD = "DictDaliApp.uid_source"
KEAR_APPLI_ISSUER_COLUMN = "KEAR_APPLI (identifiers.issuer)"
KEAR_APPLI_IDENTIFIER_COLUMN = "KEAR_APPLI (identifiers.identifier)"
PROPOSED_APPLICATION_LABEL_COLUMN = "proposed application label"
ORDERED_APPLICATION_LABEL_ATTRIBUTES = ["IRT", "IAPPLI (Trigram)", "IAPPLI"]
APPLICATION_DICTIONARY_HEADERS = [
    "uid",
    DICT_DALI_APP_SOURCE_FIELD,
    "name",
    "short_label",
    "irt_code",
    "iappli_code",
    "trigram",
    "dsi",
    "application_management_rc",
    "application_development_manager",
    "asa",
    "status",
    KEAR_APPLI_ISSUER_COLUMN,
    KEAR_APPLI_IDENTIFIER_COLUMN,
    PROPOSED_APPLICATION_LABEL_COLUMN,
]
ACCOUNT_MAPPING_SENTINELS = {"NOT_AVAILABLE", "NOT_GEN2"}
FILTER_DEFINITIONS: List[Dict[str, str]] = [
    {"name": "F_INTEXP.INCLUDE_server_os_name", "field": "server_os_name", "mode": "include_exact"},
    {"name": "F_INTEXP.INCLUDE_server_cloud_type", "field": "server_cloud_type", "mode": "include_exact"},
    {"name": "F_INTEXP.EXCLUDE_application_dali_dsi", "field": "application_dali_dsi", "mode": "exclude_contains"},
    {"name": "F_INTEXP.INCLUDE_server_status", "field": "server_status", "mode": "include_exact"},
    {"name": "F_INTEXP.EXCLUDE_server_typology", "field": "server_typology", "mode": "exclude_contains"},
    {"name": "F_INTEXP.INCLUDE_server_environment", "field": "server_environment", "mode": "include_contains"},
    {"name": "F_INTEXP.EXCLUDE_server_silo", "field": "server_silo", "mode": "exclude_contains"},
]


def setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s | %(levelname)s | %(name)s | %(message)s")


def read_filters_conf(filters_file: str) -> Dict[str, str]:
    filters: Dict[str, str] = {}
    with open(filters_file, "r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, value = line.split("=", 1)
            elif "," in line:
                key, value = line.split(",", 1)
            else:
                log.warning("Ignoring invalid filter line (expected key=value or key,value): %s", line)
                continue
            key = key.strip()
            value = value.strip()
            if key:
                filters[key] = value
    return filters


def parse_filter_tokens(filters: Dict[str, str], filter_name: str) -> List[str]:
    raw_value = filters.get(filter_name, "")
    tokens: List[str] = []
    seen = set()
    for raw_token in raw_value.split(","):
        token = raw_token.strip().strip("*").strip().casefold()
        if token and token not in seen:
            seen.add(token)
            tokens.append(token)
    return tokens


def apply_filter(value: Any, tokens: List[str], mode: str) -> str:
    if not tokens:
        return "Y"
    normalized_value = value_to_text(value).strip().casefold()
    if mode == "include_exact":
        return "Y" if normalized_value in tokens else "N"
    if mode == "include_contains":
        return "Y" if any(token in normalized_value for token in tokens) else "N"
    if mode == "exclude_contains":
        return "N" if any(token in normalized_value for token in tokens) else "Y"
    raise ValueError(f"Unsupported INTERNET.EXPOSED filter mode: {mode}")


def validate_filter_targets(source_fields: List[str]) -> None:
    missing = [definition["field"] for definition in FILTER_DEFINITIONS if definition["field"] not in source_fields]
    if missing:
        raise ValueError(f"INTERNET.EXPOSED filter target columns are missing from source fields: {', '.join(missing)}")


def build_fieldnames(source_fields: List[str]) -> List[str]:
    validate_filter_targets(source_fields)
    filters_by_field: Dict[str, List[str]] = {}
    for definition in FILTER_DEFINITIONS:
        filters_by_field.setdefault(definition["field"], []).append(definition["name"])

    fieldnames = list(TECHNICAL_FIELDS)
    for field in source_fields:
        fieldnames.append(field)
        fieldnames.extend(filters_by_field.get(field, []))
    fieldnames.extend(INVENTORY_ENRICHMENT_FIELDS)
    fieldnames.extend(PCE_WORKLOAD_FIELDS)
    fieldnames.append(ALL_FILTERS_FIELD)
    fieldnames.extend(MARLEY_KEAR_FIELDS)
    return fieldnames


def all_filter_results(row: Dict[str, Any]) -> List[str]:
    filter_results = []
    for definition in FILTER_DEFINITIONS:
        if definition["name"] == "F_INTEXP.INCLUDE_server_environment":
            filter_results.append(value_to_text(row.get(CALCULATED_ENV_FILTER_FIELD, "Y")))
        else:
            filter_results.append(value_to_text(row.get(definition["name"], "Y")))
    return filter_results


def refresh_all_filters(row: Dict[str, Any]) -> None:
    row[ALL_FILTERS_FIELD] = "Y" if all(value == "Y" for value in all_filter_results(row)) else "N"


def apply_internet_exposed_filters(rows: List[Dict[str, Any]], filters: Dict[str, str]) -> None:
    for row in rows:
        for definition in FILTER_DEFINITIONS:
            filter_name = definition["name"]
            result = apply_filter(
                row.get(definition["field"], ""),
                parse_filter_tokens(filters, filter_name),
                definition["mode"],
            )
            row[filter_name] = result
        row[CALCULATED_ENV_FILTER_FIELD] = row.get("F_INTEXP.INCLUDE_server_environment", "Y")
        refresh_all_filters(row)


def is_gen2_row(row: Dict[str, Any]) -> bool:
    return value_to_text(row.get("server_cloud_type", "")).strip().casefold() == "gen 2"


def normalize_lookup_uid(value: Any) -> str:
    return value_to_text(value).strip().upper()


def inventory_hostid_from_server_uid(value: Any) -> str:
    uid = normalize_lookup_uid(value)
    if not uid:
        return ""
    return uid if uid.startswith("VM_") else f"VM_{uid}"


def server_uid_from_inventory_hostid(value: Any) -> str:
    hostid = normalize_lookup_uid(value)
    if hostid.startswith("VM_"):
        return hostid[3:]
    return hostid


def missing_enrichment_value(row: Dict[str, Any]) -> str:
    return "NOT_AVAILABLE" if is_gen2_row(row) else "NOT_GEN2"


def normalize_enrichment_value(value: Any, fallback: str) -> str:
    text = value_to_text(value).strip()
    return text if text else fallback


def apply_inventory_enrichment(rows: List[Dict[str, Any]], inventory_by_uid: Dict[str, Dict[str, Any]]) -> None:
    for row in rows:
        fallback = missing_enrichment_value(row)
        for field in INVENTORY_ENRICHMENT_FIELDS:
            row[field] = fallback
        if not is_gen2_row(row):
            continue
        uid = normalize_lookup_uid(row.get("server_uid", ""))
        inventory_row = inventory_by_uid.get(uid, {})
        for field, source_field in (
            ("INV_owner_app_name", "owner_app_name"),
            ("INV_beneficiary", "beneficiary"),
            ("INV_region", "region"),
        ):
            row[field] = normalize_enrichment_value(inventory_row.get(source_field, ""), fallback)


def apply_platform_account_mapping(rows: List[Dict[str, Any]], dict_account_rows: List[Dict[str, str]]) -> None:
    accounts_by_key = {
        normalize_account_key(row.get("account", "")): row
        for row in dict_account_rows
        if normalize_account_key(row.get("account", ""))
    }
    for row in rows:
        fallback = missing_enrichment_value(row)
        owner = value_to_text(row.get("INV_owner_app_name", "")).strip()
        beneficiary = value_to_text(row.get("INV_beneficiary", "")).strip()
        owner_key = normalize_account_key(owner)
        beneficiary_key = normalize_account_key(beneficiary)
        owner_account = accounts_by_key.get(owner_key, {})
        beneficiary_account = accounts_by_key.get(beneficiary_key, {})
        row["PA_owner_id"] = (
            owner
            if owner_key in ACCOUNT_MAPPING_SENTINELS
            else normalize_enrichment_value(owner_account.get("id", ""), fallback)
        )
        row["PA_beneficiary_id"] = (
            beneficiary
            if beneficiary_key in ACCOUNT_MAPPING_SENTINELS
            else normalize_enrichment_value(beneficiary_account.get("id", ""), fallback)
        )
        row["PA_beneficiary_ENV"] = (
            beneficiary
            if beneficiary_key in ACCOUNT_MAPPING_SENTINELS
            else normalize_enrichment_value(beneficiary_account.get("env", ""), fallback)
        )



def normalize_column_key(value: Any) -> str:
    return "".join(ch for ch in value_to_text(value).casefold() if ch.isalnum())


def row_value_by_candidates(row: Dict[str, Any], candidates: List[str]) -> str:
    wanted = {normalize_column_key(candidate) for candidate in candidates}
    for key, value in row.items():
        if normalize_column_key(key) in wanted:
            return value_to_text(value).strip()
    return ""


def short_hostname(value: Any) -> str:
    raw = value_to_text(value).strip()
    return raw.split(".", 1)[0].strip() if raw else ""


def is_ip_derived_name(value: str) -> bool:
    if not _IP_DERIVED_NAME_RE.fullmatch(value):
        return False
    ip_slug = value[3:] if value.upper().startswith("IP-") else value
    parts = ip_slug.split("-")
    return len(parts) == 4 and all(part.isdigit() and 0 <= int(part) <= 255 for part in parts)


def expand_ip_derived_name_variants(value: Any) -> List[str]:
    normalized = normalize_lookup_uid(value)
    if not normalized:
        return []
    if not is_ip_derived_name(normalized):
        return [normalized]
    without_prefix = normalized[3:] if normalized.startswith("IP-") else normalized
    with_prefix = f"IP-{without_prefix}"
    return [normalized, without_prefix] if normalized.startswith("IP-") else [normalized, with_prefix]


def parse_managed_flag(value: Any) -> bool:
    return normalize_lookup_uid(value) in {"TRUE", "1", "YES", "Y"}


def read_pce_workload_rows(workload_csv: Path) -> List[Dict[str, str]]:
    if not workload_csv.is_file():
        log.warning("PCE workload derived CSV not found: %s", workload_csv)
        return []
    with workload_csv.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [{key: value_to_text(value).strip() for key, value in row.items()} for row in reader]
    log.info("PCE workload derived CSV loaded rows=%s path=%s", len(rows), workload_csv)
    return rows


def build_pce_workload_indexes(
    workload_rows: List[Dict[str, str]],
) -> Tuple[
    Dict[str, Dict[str, str]],
    Dict[str, Dict[str, str]],
    Dict[str, Dict[str, str]],
    Dict[str, Dict[str, str]],
    Dict[str, Dict[str, str]],
]:
    by_external_ref: Dict[str, Dict[str, str]] = {}
    managed_short: Dict[str, Dict[str, str]] = {}
    managed_ocs: Dict[str, Dict[str, str]] = {}
    unmanaged_short: Dict[str, Dict[str, str]] = {}
    unmanaged_ocs: Dict[str, Dict[str, str]] = {}

    def add(index: Dict[str, Dict[str, str]], key: str, row: Dict[str, str]) -> None:
        normalized = normalize_lookup_uid(key)
        if normalized and normalized not in index:
            index[normalized] = row

    for row in workload_rows:
        add(by_external_ref, row_value_by_candidates(row, ["external_data_reference"]), row)
        target_short = managed_short if parse_managed_flag(row_value_by_candidates(row, ["managed"])) else unmanaged_short
        target_ocs = managed_ocs if parse_managed_flag(row_value_by_candidates(row, ["managed"])) else unmanaged_ocs
        add(target_short, short_hostname(row_value_by_candidates(row, ["short_hostname"])), row)
        for variant in expand_ip_derived_name_variants(row_value_by_candidates(row, ["ocs_name_from_IP"])):
            add(target_ocs, variant, row)
    return by_external_ref, managed_short, managed_ocs, unmanaged_short, unmanaged_ocs


def pce_lookup_candidates(row: Dict[str, Any]) -> List[str]:
    candidates: List[str] = []
    for field in ("server_hostname", "server_name"):
        raw = value_to_text(row.get(field, "")).strip()
        for candidate in (raw, short_hostname(raw)):
            normalized = normalize_lookup_uid(candidate)
            if normalized and normalized not in candidates:
                candidates.append(normalized)
    return candidates


def apply_pce_workload_enrichment(rows: List[Dict[str, Any]], workload_rows: List[Dict[str, str]]) -> None:
    for row in rows:
        for field in PCE_WORKLOAD_FIELDS:
            row[field] = ""
        row["PCE_match_status"] = "NOT_FOUND"
    if not workload_rows:
        return

    by_external_ref, managed_short, managed_ocs, unmanaged_short, unmanaged_ocs = build_pce_workload_indexes(workload_rows)
    for row in rows:
        match = None
        method = ""
        server_uid = normalize_lookup_uid(row.get("server_uid", ""))
        if server_uid:
            match = by_external_ref.get(server_uid)
            if match:
                method = "external_data_reference=server_uid"
        if not match:
            for candidate in pce_lookup_candidates(row):
                match = managed_short.get(candidate)
                if match:
                    method = "managed short_hostname fallback"
                    break
                match = managed_ocs.get(candidate)
                if match:
                    method = "managed ocs_name_from_IP fallback"
                    break
                match = unmanaged_short.get(candidate)
                if match:
                    method = "unmanaged short_hostname fallback"
                    break
                match = unmanaged_ocs.get(candidate)
                if match:
                    method = "unmanaged ocs_name_from_IP fallback"
                    break
        if not match:
            continue
        managed = parse_managed_flag(row_value_by_candidates(match, ["managed"]))
        row["PCE_match_status"] = "MANAGED_WORKLOAD" if managed else "UNMANAGED_WORKLOAD"
        row["PCE_match_method"] = method
        for target, candidates in PCE_WORKLOAD_SOURCE_FIELDS.items():
            row[target] = row_value_by_candidates(match, candidates)

def apply_calculated_environment_filter(rows: List[Dict[str, Any]], filters: Dict[str, str]) -> None:
    tokens = parse_filter_tokens(filters, "F_INTEXP.INCLUDE_server_environment")
    for row in rows:
        if is_gen2_row(row):
            value = value_to_text(row.get("PA_beneficiary_ENV", "")).strip().casefold()
            row[CALCULATED_ENV_FILTER_FIELD] = "Y" if not tokens or value in tokens else "N"
        else:
            row[CALCULATED_ENV_FILTER_FIELD] = apply_filter(
                row.get("server_environment", ""),
                tokens,
                "include_contains",
            )
        refresh_all_filters(row)


def fetch_inventory_enrichment(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    lookup_hostids = sorted(
        {inventory_hostid_from_server_uid(row.get("server_uid", "")) for row in rows if is_gen2_row(row) and inventory_hostid_from_server_uid(row.get("server_uid", ""))}
    )
    if not lookup_hostids:
        log.info("Data4Sec inventory enrichment skipped: no Gen 2 server_uid values")
        return {}

    cfg = QUERY_CONFIG["inventory"]
    source_fields = ["hostid", "owner_app_name", "beneficiary", "region"]
    client = Data4secClient()
    if not client.es_connection:
        raise RuntimeError("No Elasticsearch connection available for Data4Sec inventory enrichment")

    log.info("Data4Sec inventory enrichment start lookup_hostids=%s", len(lookup_hostids))
    result_map = client.bulk_search_multi(
        index_name=cfg["index"],
        search_field="hostid",
        values=lookup_hostids,
        source_fields=source_fields,
        scroll_timeout=QUERY_CONFIG.get("scroll_timeout", "10m"),
        size=QUERY_CONFIG.get("batch_size", 500),
        term_filters=None,
    )

    output: Dict[str, Dict[str, Any]] = {}
    for hostid, docs in result_map.items():
        if docs:
            output[server_uid_from_inventory_hostid(hostid)] = docs[0]
    log.info("Data4Sec inventory enrichment done lookup_hostids=%s matched=%s", len(lookup_hostids), len(output))
    return output


def normalize_platform_tag_key(value: Any) -> str:
    return "".join(ch for ch in value_to_text(value).upper() if ch.isalnum())


def extract_platform_tag_value(tags_value: Any, tag_names: Set[str]) -> str:
    wanted = {normalize_platform_tag_key(name) for name in tag_names}
    pairs: List[tuple] = []
    if isinstance(tags_value, dict):
        pairs.extend((value_to_text(key), value_to_text(value)) for key, value in tags_value.items())
    elif isinstance(tags_value, list):
        for item in tags_value:
            if isinstance(item, dict):
                pairs.extend((value_to_text(key), value_to_text(value)) for key, value in item.items())
            else:
                text = value_to_text(item).strip()
                if ":" in text:
                    key, value = text.split(":", 1)
                    pairs.append((key, value))
    else:
        text = value_to_text(tags_value).strip()
        if text.startswith("[") and text.endswith("]"):
            text = text[1:-1]
        for part in text.split(","):
            if ":" in part:
                key, value = part.split(":", 1)
                pairs.append((key, value))

    for key, value in pairs:
        if normalize_platform_tag_key(key) in wanted:
            return value_to_text(value).strip()
    return ""


def normalize_account_key(value: Any) -> str:
    return value_to_text(value).strip().upper()


def distinct_inventory_accounts(rows: List[Dict[str, Any]]) -> List[str]:
    accounts_by_key: Dict[str, str] = {}
    for row in rows:
        for field in ("INV_owner_app_name", "INV_beneficiary"):
            account = value_to_text(row.get(field, "")).strip()
            key = normalize_account_key(account)
            if key and key not in ACCOUNT_MAPPING_SENTINELS and key not in accounts_by_key:
                accounts_by_key[key] = account
    return sorted(accounts_by_key.values(), key=lambda item: item.casefold())


def fetch_platform_account_dictionary(accounts: List[str]) -> List[Dict[str, str]]:
    account_by_key = {normalize_account_key(account): account for account in accounts if normalize_account_key(account)}
    if not account_by_key:
        log.info("Platform account dictionary skipped: no owner_app_name/beneficiary values")
        return []

    cfg = QUERY_CONFIG.get("platform_accounts", {})
    index_name = str(cfg.get("index", "platform_accounts"))
    search_field = str(cfg.get("search_field", "name"))
    source_fields = sorted(set(list(cfg.get("source_fields", ["name", "tags"])) + ["id", "tags"]))
    client = Data4secClient()
    if not client.es_connection:
        raise RuntimeError("No Elasticsearch connection available for Data4Sec platform_accounts dictionary")

    lookup_values = sorted(account_by_key.keys())
    log.info("Platform account dictionary lookup start accounts=%s", len(lookup_values))
    result_map = client.bulk_search_multi(
        index_name=index_name,
        search_field=search_field,
        values=lookup_values,
        source_fields=source_fields,
        scroll_timeout=QUERY_CONFIG.get("scroll_timeout", "10m"),
        size=QUERY_CONFIG.get("batch_size", 500),
        term_filters=cfg.get("term_filters", {}),
    )

    dictionary_rows: List[Dict[str, str]] = []
    for account_key in lookup_values:
        account = account_by_key[account_key]
        docs = result_map.get(account_key, [])
        account_id = ""
        env = ""
        for doc in docs:
            tags = doc.get("tags", "")
            account_id = extract_platform_tag_value(tags, {"ID", "ACCOUNT_ID", "ACCOUNTID", "APP_ID", "APPID"}) or value_to_text(doc.get("id", ""))
            env = extract_platform_tag_value(tags, {"ENV", "ENVIRONMENT"})
            if account_id or env:
                break
        dictionary_rows.append({"account": account, "id": account_id, "env": env})
    log.info("Platform account dictionary lookup done accounts=%s matched=%s", len(lookup_values), sum(1 for row in dictionary_rows if row.get("id") or row.get("env")))
    return dictionary_rows


def build_application_search_body(uid: str, limit: int = 100) -> Dict[str, Any]:
    return {
        "filters": [
            {
                "attributeName": "uid",
                "attributeValue": uid,
                "matchType": "equals",
            }
        ],
        "includeCount": True,
        "label": "Application",
        "limit": limit,
        "orderBy": [{"direction": "asc", "labelProperty": "string"}],
        "skip": 0,
    }


def extract_application_properties(response: Dict[str, Any]) -> Dict[str, Any]:
    result = response.get("result") if isinstance(response, dict) else None
    if not isinstance(result, list) or not result:
        return {}
    first = result[0]
    if not isinstance(first, dict):
        return {}
    leading_node = first.get("leading_node")
    if not isinstance(leading_node, dict):
        return {}
    properties = leading_node.get("properties")
    return properties if isinstance(properties, dict) else {}


def collect_dict_dali_app_uids(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    collected: List[Dict[str, str]] = []
    seen: Set[str] = set()
    for row in rows:
        if value_to_text(row.get(ALL_FILTERS_FIELD, "")).strip().upper() != "Y":
            continue
        calculated_kear = value_to_text(row.get(CALCULATED_SINGLE_KEAR_FIELD, "")).strip()
        if calculated_kear and calculated_kear not in {"MULTIPLE_KEARS", MISSING_KEAR_VALUE}:
            candidates = [(calculated_kear, CALCULATED_SINGLE_KEAR_FIELD)]
        elif calculated_kear == "MULTIPLE_KEARS":
            candidates = [(uid, "MULTIPLE_KEARS") for uid in split_application_uids(row.get(MARLEY_KEAR_UUID_FIELD, ""))]
        else:
            candidates = []
        for uid, source in candidates:
            key = uid.strip()
            if not key or key in seen:
                continue
            seen.add(key)
            collected.append({"uid": key, DICT_DALI_APP_SOURCE_FIELD: source})
    return collected


def build_dict_dali_app_rows(
    client: Any,
    uid_rows: List[Dict[str, str]],
    search_endpoint: str,
    limit: int = 100,
) -> List[Dict[str, str]]:
    output: List[Dict[str, str]] = []
    for index, uid_row in enumerate(uid_rows, start=1):
        uid = uid_row["uid"]
        source = uid_row.get(DICT_DALI_APP_SOURCE_FIELD, "")
        log.info("DictDaliApp DALI application lookup uid=%s progress=%s/%s source=%s", uid, index, len(uid_rows), source)
        request_body = build_application_search_body(uid, limit=limit)
        if hasattr(client, "post_json"):
            response = client.post_json(endpoint=search_endpoint, payload=request_body)
        else:
            response = dali_search_post_json(client, endpoint=search_endpoint, payload=request_body)
        properties = extract_application_properties(response)
        row = {header: value_to_text(properties.get(header, "")) for header in APPLICATION_DICTIONARY_HEADERS}
        row["uid"] = row.get("uid") or uid
        row[DICT_DALI_APP_SOURCE_FIELD] = source
        output.append(row)
    return output


def dali_search_post_json(client: DaliImpactAnalysisClient, endpoint: str, payload: Dict[str, Any], timeout_s: int = 60, retries: int = 4) -> Dict[str, Any]:
    import requests

    url = urljoin(f"{client.base_url}/", endpoint.lstrip("/"))
    filters = payload.get("filters") if isinstance(payload, dict) else None
    uid = ""
    if isinstance(filters, list):
        for item in filters:
            if isinstance(item, dict) and item.get("attributeName") == "uid":
                uid = str(item.get("attributeValue") or "")
                break

    last_exc: Optional[Exception] = None
    for attempt in range(retries + 1):
        try:
            headers = client.dali_headers(force_refresh=attempt > 0)
            headers["Content-Type"] = "application/json"
            response = requests.post(url, json=payload, headers=headers, timeout=timeout_s, verify=client.verify)
            status_code = int(response.status_code)
            if status_code in {401, 403}:
                client._token = None
                client._token_expiry_epoch = 0
                if attempt < retries:
                    continue
            if status_code in {429, 500, 502, 503, 504} and attempt < retries:
                delay = (2**attempt) + random.uniform(0, 0.5)
                log.warning("DALI search transient status=%s for uid=%s, retry in %.2fs", status_code, uid, delay)
                time.sleep(delay)
                continue
            if status_code >= 400:
                compact_body = " ".join(str(response.text or "").split())
                raise RuntimeError(f"DALI search request failed for uid={uid}: HTTP {status_code} | response={compact_body[:500]}")
            response.raise_for_status()
            return response.json()
        except requests.RequestException as exc:
            last_exc = exc
            if attempt < retries:
                delay = (2**attempt) + random.uniform(0, 0.5)
                log.warning("DALI search request error for uid=%s on attempt %s/%s: %s; retry in %.2fs", uid, attempt + 1, retries + 1, exc, delay)
                time.sleep(delay)
                continue
            break
    raise RuntimeError(f"DALI search request failed after retries for uid={uid}: {last_exc}")


def normalize_kear_appli_lookup_value(value: Any) -> str:
    return value_to_text(value).strip().upper()


def build_apma_value(issuers: List[str], identifiers: List[str]) -> str:
    mapping = dict(zip(issuers, identifiers))
    result_values: List[str] = []
    for attribute in ORDERED_APPLICATION_LABEL_ATTRIBUTES:
        value = value_to_text(mapping.get(attribute, "")).strip()
        if value:
            result_values.append(value)
    return ".".join(result_values)


def build_proposed_application_label(global_id: str, issuers: List[str], identifiers: List[str]) -> str:
    normalized_global_id = value_to_text(global_id).strip()
    concatenated = build_apma_value(issuers, identifiers)
    return f"APMA_{normalized_global_id}_{concatenated}" if concatenated else f"APMA_{normalized_global_id}"


def list_values(value: Any) -> List[Any]:
    if value is None:
        return []
    return value if isinstance(value, list) else [value]


def extract_identifier_pairs(doc: Dict[str, Any]) -> Tuple[List[str], List[str]]:
    issuers: List[str] = []
    identifiers: List[str] = []

    nested_identifiers = doc.get("identifiers")
    if isinstance(nested_identifiers, list):
        for item in nested_identifiers:
            if not isinstance(item, dict):
                continue
            issuer = value_to_text(item.get("issuer")).strip()
            identifier = value_to_text(item.get("identifier")).strip()
            if issuer or identifier:
                issuers.append(issuer)
                identifiers.append(identifier)
    elif isinstance(nested_identifiers, dict):
        issuer = value_to_text(nested_identifiers.get("issuer")).strip()
        identifier = value_to_text(nested_identifiers.get("identifier")).strip()
        if issuer or identifier:
            issuers.append(issuer)
            identifiers.append(identifier)

    if issuers or identifiers:
        return issuers, identifiers

    dotted_issuers = [value_to_text(value).strip() for value in list_values(doc.get("identifiers.issuer"))]
    dotted_identifiers = [value_to_text(value).strip() for value in list_values(doc.get("identifiers.identifier"))]
    max_len = max(len(dotted_issuers), len(dotted_identifiers))
    for index in range(max_len):
        issuer = dotted_issuers[index] if index < len(dotted_issuers) else ""
        identifier = dotted_identifiers[index] if index < len(dotted_identifiers) else ""
        if issuer or identifier:
            issuers.append(issuer)
            identifiers.append(identifier)
    return issuers, identifiers


def query_kear_appli_by_global_ids(uids: List[str]) -> Dict[str, Dict[str, Any]]:
    lookup_values_by_key: Dict[str, str] = {}
    for uid in uids:
        raw_uid = value_to_text(uid).strip()
        normalized = normalize_kear_appli_lookup_value(raw_uid)
        if raw_uid and normalized not in lookup_values_by_key:
            lookup_values_by_key[normalized] = raw_uid
    if not lookup_values_by_key:
        log.info("DictDaliApp KEAR_APPLI enrichment skipped: no uid to query")
        return {}

    client = Data4secClient()
    if not client.es_connection:
        raise RuntimeError("No Elasticsearch connection available for Data4Sec kear_appli enrichment")

    lookup_values = list(lookup_values_by_key.keys())
    index_name = (os.getenv("KEAR_APPLI_INDEX") or "kear_appli").strip().strip("'").strip('"')
    search_field = (os.getenv("KEAR_APPLI_SEARCH_FIELD") or "global_id").strip().strip("'").strip('"')
    scroll_timeout = (os.getenv("KEAR_APPLI_SCROLL_TIMEOUT") or QUERY_CONFIG.get("scroll_timeout", "10m")).strip().strip("'").strip('"')
    batch_size = int((os.getenv("KEAR_APPLI_BATCH_SIZE") or str(QUERY_CONFIG.get("batch_size", 500))).strip().strip("'").strip('"'))
    source_fields = ["global_id", "identifiers", "identifiers.issuer", "identifiers.identifier"]
    log.info("DictDaliApp KEAR_APPLI enrichment query start index=%s global_id_count=%s", index_name, len(lookup_values))
    docs_by_uid = client.bulk_search_multi(
        index_name=index_name,
        search_field=search_field,
        values=lookup_values,
        source_fields=source_fields,
        scroll_timeout=scroll_timeout,
        size=batch_size,
    )

    docs_by_normalized_uid: Dict[str, Dict[str, Any]] = {}
    for lookup_uid, docs in docs_by_uid.items():
        if docs:
            docs_by_normalized_uid[normalize_kear_appli_lookup_value(lookup_uid)] = docs[0]
    log.info("DictDaliApp KEAR_APPLI enrichment query done matched_global_ids=%s/%s", len(docs_by_normalized_uid), len(lookup_values))
    return docs_by_normalized_uid


def enrich_dict_dali_app_rows_with_kear_appli(rows: List[Dict[str, str]]) -> None:
    for row in rows:
        row.setdefault(KEAR_APPLI_ISSUER_COLUMN, "")
        row.setdefault(KEAR_APPLI_IDENTIFIER_COLUMN, "")
        row.setdefault(PROPOSED_APPLICATION_LABEL_COLUMN, "")

    docs_by_uid = query_kear_appli_by_global_ids([row.get("uid", "") for row in rows])
    matched_rows = 0
    for row in rows:
        uid = value_to_text(row.get("uid")).strip()
        doc = docs_by_uid.get(normalize_kear_appli_lookup_value(uid))
        if not doc:
            continue
        global_id = value_to_text(doc.get("global_id")).strip() or uid
        issuers, identifiers = extract_identifier_pairs(doc)
        row[KEAR_APPLI_ISSUER_COLUMN] = ", ".join(issuers)
        row[KEAR_APPLI_IDENTIFIER_COLUMN] = ", ".join(identifiers)
        row[PROPOSED_APPLICATION_LABEL_COLUMN] = build_proposed_application_label(global_id, issuers, identifiers)
        matched_rows += 1
    log.info("DictDaliApp KEAR_APPLI enrichment done rows=%s matched_rows=%s", len(rows), matched_rows)


def fetch_dict_dali_app_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    uid_rows = collect_dict_dali_app_uids(rows)
    if not uid_rows:
        log.info("DictDaliApp skipped: no application UID from F_ALL_FILTERS=Y rows")
        return []
    search_endpoint = (os.getenv("DALI_SEARCH_ENDPOINT") or "/api/v1/search").strip().strip("'").strip('"')
    client = DaliImpactAnalysisClient()
    dict_rows = build_dict_dali_app_rows(client, uid_rows, search_endpoint=str(search_endpoint))
    enrich_dict_dali_app_rows_with_kear_appli(dict_rows)
    return dict_rows





def split_application_uids(value: Any) -> List[str]:
    tokens: List[str] = []
    seen: Set[str] = set()
    for token in value_to_text(value).split(","):
        normalized = token.strip()
        if normalized and normalized not in seen:
            seen.add(normalized)
            tokens.append(normalized)
    return tokens


def nested_values(value: Any, path: str) -> List[Any]:
    parts = path.split(".") if path else []

    def walk(current: Any, remaining: List[str]) -> Iterable[Any]:
        if not remaining:
            if isinstance(current, list):
                for item in current:
                    yield item
            else:
                yield current
            return
        if isinstance(current, list):
            for item in current:
                yield from walk(item, remaining)
        elif isinstance(current, dict):
            yield from walk(current.get(remaining[0]), remaining[1:])

    return [item for item in walk(value, parts) if value_to_text(item).strip()]


def format_marley_values(values: List[Any], deduplicate: bool = True) -> str:
    output: List[str] = []
    seen: Set[str] = set()
    for value in values:
        text = value_to_text(value).strip()
        if not text:
            continue
        if deduplicate and text in seen:
            continue
        seen.add(text)
        output.append(text)
    return ", ".join(output)


def format_factor_value(value: Any) -> str:
    factor = parse_factor(value)
    if factor is None:
        return value_to_text(value).strip()
    if factor.is_integer():
        return str(int(factor))
    return str(factor)


def parse_factor(value: Any) -> Optional[float]:
    text = value_to_text(value).strip().replace("%", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def app_info_pairs(doc: Dict[str, Any]) -> List[Tuple[str, Optional[float]]]:
    app_info = doc.get("app_info", {})
    if isinstance(app_info, list):
        pairs: List[Tuple[str, Optional[float]]] = []
        for item in app_info:
            if not isinstance(item, dict):
                continue
            factor = item.get("kear_factor", item.get("factor"))
            for uuid in nested_values(item, "kear_uuid"):
                pairs.append((value_to_text(uuid).strip(), parse_factor(factor)))
        return [(uuid, factor) for uuid, factor in pairs if uuid]
    uuids = nested_values(doc, "app_info.kear_uuid")
    factors = nested_values(doc, "app_info.kear_factor") or nested_values(doc, "app_info.factor")
    pairs = []
    for index, uuid in enumerate(uuids):
        factor = parse_factor(factors[index]) if index < len(factors) else None
        pairs.append((value_to_text(uuid).strip(), factor))
    return [(uuid, factor) for uuid, factor in pairs if uuid]


def calculate_single_kear(pairs: List[Tuple[str, Optional[float]]]) -> str:
    totals: Dict[str, float] = {}
    for uuid, factor in pairs:
        totals[uuid] = totals.get(uuid, 0.0) + (factor if factor is not None else 0.0)
    if not totals:
        return ""
    max_factor = max(totals.values())
    winners = [uuid for uuid, total in totals.items() if total == max_factor]
    return winners[0] if len(winners) == 1 else "MULTIPLE_KEARS"


def fetch_marley_kear_by_server_uid(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    lookup_uids = sorted({
        normalize_lookup_uid(row.get("server_uid", ""))
        for row in rows
        if value_to_text(row.get(ALL_FILTERS_FIELD, "")).strip().upper() == "Y"
        and len(split_application_uids(row.get("application_uid", ""))) != 1
        and normalize_lookup_uid(row.get("server_uid", ""))
    })
    if not lookup_uids:
        log.info("Data4Sec marley_original kear enrichment skipped: no F_ALL_FILTERS=Y rows with empty or multiple application_uid values")
        return {}
    cfg = QUERY_CONFIG.get("marley_original", {})
    client = Data4secClient()
    if not client.es_connection:
        raise RuntimeError("No Elasticsearch connection available for Data4Sec marley_original kear enrichment")
    source_fields = sorted(set(list(cfg.get("source_fields", [])) + ["uuid", "app_info.kear_uuid", "app_info.kear_factor", "app_info.factor", "app_info"]))
    log.info("Data4Sec marley_original kear enrichment start lookup_server_uids=%s", len(lookup_uids))
    return client.bulk_search_multi(
        index_name=str(cfg.get("index", "marley_original")),
        search_field="uuid",
        values=lookup_uids,
        source_fields=source_fields,
        scroll_timeout=QUERY_CONFIG.get("scroll_timeout", "10m"),
        size=QUERY_CONFIG.get("batch_size", 500),
        term_filters=cfg.get("term_filters", {}),
    )


def apply_marley_kear_enrichment(rows: List[Dict[str, Any]], marley_docs_by_uid: Dict[str, List[Dict[str, Any]]]) -> None:
    for row in rows:
        for field in MARLEY_KEAR_FIELDS:
            row[field] = ""
        app_uids = split_application_uids(row.get("application_uid", ""))
        if len(app_uids) == 1:
            row[CALCULATED_SINGLE_KEAR_FIELD] = app_uids[0]
            continue
        if value_to_text(row.get(ALL_FILTERS_FIELD, "")).strip().upper() != "Y":
            continue
        uid = normalize_lookup_uid(row.get("server_uid", ""))
        docs = marley_docs_by_uid.get(uid, [])
        pairs: List[Tuple[str, Optional[float]]] = []
        for doc in docs:
            pairs.extend(app_info_pairs(doc))
        row[MARLEY_KEAR_UUID_FIELD] = format_marley_values([uuid for uuid, _factor in pairs])
        row[MARLEY_KEAR_FACTOR_FIELD] = format_marley_values(
            [format_factor_value(factor) for _uuid, factor in pairs if factor is not None],
            deduplicate=False,
        )
        row[CALCULATED_SINGLE_KEAR_FIELD] = calculate_single_kear(pairs) or MISSING_KEAR_VALUE


def build_internet_exposed_query(cfg: Dict[str, Any], size: int) -> Dict[str, Any]:
    return {
        "_source": cfg["source_fields"],
        "query": {
            "bool": {
                "filter": [{"terms": {cfg["usage_field"]: cfg["usage_values"]}}],
                "should": [
                    {"terms": {cfg["dali_exposed_field"]: cfg["dali_exposed_values"]}},
                    {
                        "wildcard": {
                            cfg["masai_exposed_field"]: {
                                "value": cfg["masai_exposed_wildcard"],
                                "case_insensitive": True,
                            }
                        }
                    },
                ],
                "minimum_should_match": 1,
            }
        },
        "size": size,
        "sort": ["_doc"],
    }


def value_to_text(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value if str(item).strip())
    if value is None:
        return ""
    return str(value)


def annotate_scope(source: Dict[str, Any], cfg: Dict[str, Any]) -> Dict[str, Any]:
    server_exposed = str(source.get("server_exposed") or "").strip().lower()
    masai_exposition = str(source.get("application_internet_exposition_masai") or "").strip().lower()
    is_dali = server_exposed == "yes"
    is_masai = "internet" in masai_exposition
    scopes = []
    if is_dali:
        scopes.append("DALI.EXPOSED")
    if is_masai:
        scopes.append("MASAI.EXPOSED")
    return {
        **{field: value_to_text(source.get(field, "")) for field in cfg["source_fields"]},
        "exposure_scopes": ", ".join(scopes),
        "is_dali_exposed": "Y" if is_dali else "N",
        "is_masai_exposed": "Y" if is_masai else "N",
    }


def dedupe_key(row: Dict[str, Any]) -> str:
    for field in ("server_uid", "server_hostname", "server_name", "server_friendly_name"):
        value = str(row.get(field) or "").strip().upper()
        if value:
            return f"{field}:{value}"
    return json.dumps(row, sort_keys=True, ensure_ascii=False)


def fetch_internet_exposed() -> List[Dict[str, Any]]:
    cfg = QUERY_CONFIG["internet_exposed"]
    client = Data4secClient()
    if not client.es_connection:
        raise RuntimeError("No Elasticsearch connection available for Data4Sec internet exposed extract")

    query = build_internet_exposed_query(cfg, size=QUERY_CONFIG.get("batch_size", 500))
    log.info("Data4Sec INTERNET.EXPOSED query index=%s payload=%s", cfg["index"], json.dumps(query, ensure_ascii=False))

    rows: List[Dict[str, Any]] = []
    seen = set()
    for hit in scan(
        client.es_connection,
        index=cfg["index"],
        query=query,
        scroll=QUERY_CONFIG.get("scroll_timeout", "10m"),
        size=QUERY_CONFIG.get("batch_size", 500),
    ):
        row = annotate_scope(hit.get("_source", {}) or {}, cfg)
        key = dedupe_key(row)
        if key not in seen:
            seen.add(key)
            rows.append(row)
    log.info("Data4Sec INTERNET.EXPOSED extract done rows=%s", len(rows))
    return rows


def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json_gz(path: Path, rows: List[Dict[str, Any]], query: Dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"meta": {"row_count": len(rows), "query": query}, "rows": rows}
    gz_path = path if path.suffix == ".gz" else path.with_suffix(path.suffix + ".gz")
    with gzip.open(gz_path, "wt", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
    return gz_path


def write_xlsx(
    path: Path,
    rows: List[Dict[str, Any]],
    fieldnames: List[str],
    dict_account_rows: Optional[List[Dict[str, str]]] = None,
    dict_dali_app_rows: Optional[List[Dict[str, str]]] = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    filter_fieldnames = {definition["name"] for definition in FILTER_DEFINITIONS}
    filter_fieldnames.add(ALL_FILTERS_FIELD)
    pce_fieldnames = set(PCE_WORKLOAD_FIELDS)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "RAW_INTERNET_EXPOSED"
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    dict_header_fill = PatternFill("solid", fgColor="8064A2")
    filter_header_fill = PatternFill("solid", fgColor="595959")
    pce_header_fill = PatternFill("solid", fgColor="F4B183")
    pce_fill = PatternFill("solid", fgColor="FCE4D6")
    filter_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True, color="FFFFFF")
    filter_columns = [idx for idx, field in enumerate(fieldnames, start=1) if field in filter_fieldnames]
    pce_columns = [idx for idx, field in enumerate(fieldnames, start=1) if field in pce_fieldnames]
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        if cell.value in pce_fieldnames:
            cell.fill = pce_header_fill
        elif cell.value in filter_fieldnames:
            cell.fill = filter_header_fill
        else:
            cell.fill = header_fill
        cell.font = header_font
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in filter_columns:
            row[col_idx - 1].fill = filter_fill
        for col_idx in pce_columns:
            row[col_idx - 1].fill = pce_fill
    for worksheet in (ws,):
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
        for column_cells in worksheet.columns:
            max_length = max(len(value_to_text(cell.value)) for cell in column_cells)
            adjusted_width = min(max(max_length + 2, 10), 60)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    stats = wb.create_sheet("STATS")
    stats.append(["metric", "value"])
    stats.append(["total_servers", len(rows)])
    stats.append(["dali_exposed_servers", sum(1 for r in rows if r.get("is_dali_exposed") == "Y")])
    stats.append(["masai_exposed_servers", sum(1 for r in rows if r.get("is_masai_exposed") == "Y")])
    stats.append(["distinct_application_uid", len({r.get("application_uid") for r in rows if str(r.get("application_uid") or "").strip()})])
    for cell in stats[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in stats.iter_rows():
        for cell in row:
            cell.border = thin_border
    for column_cells in stats.columns:
        max_length = max(len(value_to_text(cell.value)) for cell in column_cells)
        adjusted_width = min(max(max_length + 2, 10), 60)
        stats.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    dict_ws = wb.create_sheet("DictAccount")
    dict_ws.append(DICT_ACCOUNT_HEADERS)
    for dict_row in dict_account_rows or []:
        dict_ws.append([dict_row.get(header, "") for header in DICT_ACCOUNT_HEADERS])
    for cell in dict_ws[1]:
        cell.fill = dict_header_fill
        cell.font = header_font
    dict_ws.freeze_panes = "A2"
    dict_ws.auto_filter.ref = dict_ws.dimensions
    for row in dict_ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    for column_cells in dict_ws.columns:
        max_length = max(len(value_to_text(cell.value)) for cell in column_cells)
        adjusted_width = min(max(max_length + 2, 14), 60)
        dict_ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    dali_app_ws = wb.create_sheet(DICT_DALI_APP_SHEET)
    dali_app_ws.append(APPLICATION_DICTIONARY_HEADERS)
    for app_row in dict_dali_app_rows or []:
        dali_app_ws.append([app_row.get(header, "") for header in APPLICATION_DICTIONARY_HEADERS])
    for cell in dali_app_ws[1]:
        cell.fill = dict_header_fill
        cell.font = header_font
    dali_app_ws.freeze_panes = "A2"
    dali_app_ws.auto_filter.ref = dali_app_ws.dimensions
    for row in dali_app_ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    for column_cells in dali_app_ws.columns:
        max_length = max(len(value_to_text(cell.value)) for cell in column_cells)
        adjusted_width = min(max(max_length + 2, 14), 60)
        dali_app_ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract INTERNET.EXPOSED servers from Data4Sec dali_servers.")
    parser.add_argument("--output", required=True, help="Output XLSX path")
    parser.add_argument("--csv-out", help="Optional CSV output path")
    parser.add_argument("--json-out", help="Optional JSON.GZ output path")
    parser.add_argument("--filters-file", default="user_inputs/filters.conf", help="Filters configuration file")
    parser.add_argument("--pce-workload-derived", help="Path to export_wkld.derived.csv (defaults to output directory)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose logging")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    setup_logging(args.verbose)
    cfg = QUERY_CONFIG["internet_exposed"]
    fieldnames = build_fieldnames(cfg["source_fields"])
    filters = read_filters_conf(args.filters_file)
    rows = fetch_internet_exposed()
    apply_internet_exposed_filters(rows, filters)
    apply_inventory_enrichment(rows, fetch_inventory_enrichment(rows))
    workload_csv = (
        Path(args.pce_workload_derived)
        if args.pce_workload_derived
        else Path(args.output).parent / "export_wkld.derived.csv"
    )
    apply_pce_workload_enrichment(rows, read_pce_workload_rows(workload_csv))
    dict_account_rows = fetch_platform_account_dictionary(distinct_inventory_accounts(rows))
    apply_platform_account_mapping(rows, dict_account_rows)
    apply_calculated_environment_filter(rows, filters)
    apply_marley_kear_enrichment(rows, fetch_marley_kear_by_server_uid(rows))
    dict_dali_app_rows = fetch_dict_dali_app_rows(rows)
    output = Path(args.output)
    write_xlsx(output, rows, fieldnames, dict_account_rows=dict_account_rows, dict_dali_app_rows=dict_dali_app_rows)
    if args.csv_out:
        write_csv(Path(args.csv_out), rows, fieldnames)
    if args.json_out:
        gz_path = write_json_gz(Path(args.json_out), rows, build_internet_exposed_query(cfg, QUERY_CONFIG.get("batch_size", 500)))
        print(f"JSON.GZ written to: {gz_path}")
    print(f"INTERNET.EXPOSED rows: {len(rows)}")
    print(f"XLSX written to: {output}")


if __name__ == "__main__":
    main()
