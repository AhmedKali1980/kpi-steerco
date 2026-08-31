import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "modules"))

from internet_exposed_extract import (
    ALL_FILTERS_FIELD,
    CALCULATED_ENV_FILTER_FIELD,
    FILTER_DEFINITIONS,
    INVENTORY_ENRICHMENT_FIELDS,
    INV_PA_HIGHLIGHT_FIELDS,
    PCE_APP_COMPARISON_FIELDS,
    PCE_OUTPUT_FIELDS,
    PCE_WORKLOAD_FIELDS,
    SCOPE_DALI_APP_ENRICHMENT_FIELDS,
    SCOPE_INTEXPOSED_FIELDS,
    SCOPE_INTEXPOSED_SHEET,
    STATS_INTEXPOSED_COLUMNS,
    STATS_INTEXPOSED_SHEET,
    MARLEY_KEAR_FIELDS,
    MISSING_KEAR_VALUE,
    CALCULATED_SINGLE_KEAR_FIELD,
    DICT_DALI_APP_SOURCE_FIELD,
    KEAR_APPLI_ISSUER_COLUMN,
    KEAR_APPLI_IDENTIFIER_COLUMN,
    PROPOSED_APPLICATION_LABEL_COLUMN,
    APPLICATION_DICTIONARY_HEADERS,
    apply_marley_kear_enrichment,
    apply_calculated_environment_filter,
    apply_internet_exposed_filters,
    apply_inventory_enrichment,
    apply_platform_account_mapping,
    apply_pce_app_label_comparison,
    apply_pce_workload_enrichment,
    build_dict_dali_app_rows,
    build_scope_intexposed_rows,
    build_stats_intexposed_rows,
    build_proposed_application_label,
    collect_dict_dali_app_uids,
    distinct_inventory_accounts,
    enrich_dict_dali_app_rows_with_kear_appli,
    extract_platform_tag_value,
    extract_identifier_pairs,
    build_fieldnames,
    inventory_hostid_from_server_uid,
    query_kear_appli_by_global_ids,
    parse_filter_tokens,
    read_filters_conf,
    write_xlsx,
    unique_fieldnames,
    server_uid_from_inventory_hostid,
)


class InternetExposedFilterTests(unittest.TestCase):
    def test_filter_columns_are_inserted_after_target_fields_and_all_filter_is_last(self):
        source_fields = [
            "server_os_name",
            "server_cloud_type",
            "application_dali_dsi",
            "application_uid",
            "server_status",
            "server_typology",
            "server_environment",
            "server_silo",
            "server_team_in_charge",
        ]
        fieldnames = build_fieldnames(source_fields)

        for definition in FILTER_DEFINITIONS:
            target_index = fieldnames.index(definition["field"])
            self.assertEqual(fieldnames[target_index + 1], definition["name"])
        self.assertEqual(fieldnames[-29:-22], INVENTORY_ENRICHMENT_FIELDS)
        self.assertEqual(fieldnames[fieldnames.index("INV_owner_app_name") + 1], "PA_owner_id")
        self.assertEqual(fieldnames[fieldnames.index("INV_beneficiary") + 1], "PA_beneficiary_id")
        self.assertEqual(fieldnames[fieldnames.index("PA_beneficiary_id") + 1], "PA_beneficiary_ENV")
        self.assertEqual(fieldnames[fieldnames.index("INV_region") + 1], CALCULATED_ENV_FILTER_FIELD)
        self.assertEqual(fieldnames[-22], ALL_FILTERS_FIELD)
        self.assertEqual(fieldnames[-21:-18], MARLEY_KEAR_FIELDS)
        self.assertEqual(fieldnames[-18:], PCE_OUTPUT_FIELDS)
        pce_app_index = fieldnames.index("PCE_app")
        self.assertEqual(fieldnames[pce_app_index + 1 : pce_app_index + 3], PCE_APP_COMPARISON_FIELDS)

    def test_unique_fieldnames_removes_duplicate_pce_columns_case_insensitively(self):
        fieldnames = [
            "server_uid",
            "PCE_hostname",
            " pce_hostname ",
            "PCE_app",
            "pce_app",
            "F_ALL_FILTERS",
        ]

        self.assertEqual(unique_fieldnames(fieldnames), ["server_uid", "PCE_hostname", "PCE_app", "F_ALL_FILTERS"])

    def test_build_fieldnames_deduplicates_pce_columns_and_keeps_them_last(self):
        source_fields = [
            "server_os_name",
            "server_cloud_type",
            "application_dali_dsi",
            "server_status",
            "application_uid",
            "PCE_hostname",
            " pce_short_hostname ",
            "server_typology",
            "server_environment",
            "server_silo",
            "server_team_in_charge",
            "PCE_match_status",
            "pce_managed",
        ]

        fieldnames = build_fieldnames(source_fields)

        self.assertEqual(fieldnames[-18:], PCE_OUTPUT_FIELDS)
        normalized_fieldnames = ["".join(ch for ch in field.casefold() if ch.isalnum()) for field in fieldnames]
        for pce_field in PCE_OUTPUT_FIELDS:
            normalized_pce_field = "".join(ch for ch in pce_field.casefold() if ch.isalnum())
            self.assertEqual(normalized_fieldnames.count(normalized_pce_field), 1)
        self.assertLess(fieldnames.index("calculated_Single_Kear"), fieldnames.index("PCE_match_status"))

    def test_filters_are_case_insensitive_and_support_exact_or_contains_modes(self):
        filters = {
            "F_INTEXP.INCLUDE_server_os_name": "LINUX",
            "F_INTEXP.INCLUDE_server_cloud_type": "GEN 2",
            "F_INTEXP.EXCLUDE_application_dali_dsi": "ayvens",
            "F_INTEXP.EXCLUDE_application_uid": "APP-BLOCKED",
            "F_INTEXP.INCLUDE_server_status": "active",
            "F_INTEXP.EXCLUDE_server_typology": "HyperVisor",
            "F_INTEXP.INCLUDE_server_environment": "prd",
            "F_INTEXP.EXCLUDE_server_silo": "oos",
            "F_INTEXP.EXCLUDE_server_team_in_charge": "outsourcing",
        }
        rows = [
            {
                "server_os_name": "Linux",
                "server_cloud_type": "Gen 2",
                "application_dali_dsi": "Retail",
                "application_uid": "APP-ALLOWED",
                "server_status": "Active",
                "server_typology": "Virtual Machine",
                "server_environment": "APP-PRD-EUR",
                "server_silo": "RUN",
                "server_team_in_charge": "Core Hosting",
            },
            {
                "server_os_name": "Linux",
                "server_cloud_type": "Gen 2",
                "application_dali_dsi": "Ayvens Platform",
                "application_uid": "APP-ALLOWED",
                "server_status": "Active",
                "server_typology": "Virtual Machine",
                "server_environment": "APP-PRD-EUR",
                "server_silo": "RUN",
                "server_team_in_charge": "Core Hosting",
            },
        ]

        apply_internet_exposed_filters(rows, filters)

        self.assertEqual(rows[0][ALL_FILTERS_FIELD], "Y")
        self.assertEqual(rows[1]["F_INTEXP.EXCLUDE_application_dali_dsi"], "N")
        self.assertEqual(rows[1][ALL_FILTERS_FIELD], "N")

    def test_server_team_in_charge_filter_excludes_matches_and_feeds_all_filters(self):
        filter_name = "F_INTEXP.EXCLUDE_server_team_in_charge"
        rows = [
            {"server_team_in_charge": "Core Hosting"},
            {"server_team_in_charge": "External Outsourcing Team"},
        ]

        apply_internet_exposed_filters(rows, {filter_name: "outsourcing"})

        self.assertEqual(rows[0][filter_name], "Y")
        self.assertEqual(rows[0][ALL_FILTERS_FIELD], "Y")
        self.assertEqual(rows[1][filter_name], "N")
        self.assertEqual(rows[1][ALL_FILTERS_FIELD], "N")
        self.assertEqual(
            SCOPE_INTEXPOSED_FIELDS.index(filter_name),
            SCOPE_INTEXPOSED_FIELDS.index("server_team_in_charge") + 1,
        )

    def test_application_uid_filter_keeps_hyphenated_uids_as_single_tokens(self):
        filters = {
            "F_INTEXP.EXCLUDE_application_uid": (
                "11111111-2222-333-4444-5555555, AAAAAAAA-BBBB-CCC-DDDD-EEEEEEE"
            )
        }

        tokens = parse_filter_tokens(filters, "F_INTEXP.EXCLUDE_application_uid")

        self.assertEqual(
            tokens,
            [
                "11111111-2222-333-4444-5555555",
                "aaaaaaaa-bbbb-ccc-dddd-eeeeeee",
            ],
        )

    def test_application_uid_filter_excludes_contains_and_feeds_all_filters(self):
        blocked_uid = "11111111-2222-333-4444-5555555"
        rows = [
            {
                "server_os_name": "Linux",
                "server_cloud_type": "Gen 2",
                "application_dali_dsi": "Retail",
                "application_uid": f"99999999-8888-777-6666-5555555, {blocked_uid}",
                "server_status": "Active",
                "server_typology": "Virtual Machine",
                "server_environment": "APP-PRD-EUR",
                "server_silo": "RUN",
            }
        ]

        apply_internet_exposed_filters(rows, {"F_INTEXP.EXCLUDE_application_uid": blocked_uid})

        self.assertEqual(rows[0]["F_INTEXP.EXCLUDE_application_uid"], "N")
        self.assertEqual(rows[0][ALL_FILTERS_FIELD], "N")

    def test_inventory_hostid_mapping_uses_vm_uppercase_prefix(self):
        self.assertEqual(inventory_hostid_from_server_uid("aaa-bbb-ccc"), "VM_AAA-BBB-CCC")
        self.assertEqual(inventory_hostid_from_server_uid("VM_AAA-BBB-CCC"), "VM_AAA-BBB-CCC")
        self.assertEqual(server_uid_from_inventory_hostid("VM_AAA-BBB-CCC"), "AAA-BBB-CCC")

    def test_inventory_enrichment_only_applies_to_gen2_rows_before_all_filters(self):
        rows = [
            {"server_cloud_type": "Gen 2", "server_uid": "srv-1"},
            {"server_cloud_type": "Gen 1", "server_uid": "srv-2"},
            {"server_cloud_type": "Gen 2", "server_uid": "srv-3"},
        ]
        apply_inventory_enrichment(
            rows,
            {
                "SRV-1": {"owner_app_name": "App One", "beneficiary": "BEN", "region": "EUR"},
                "SRV-2": {"owner_app_name": "App Two", "beneficiary": "BEN2", "region": "AMER"},
            },
        )

        self.assertEqual(rows[0]["INV_owner_app_name"], "App One")
        self.assertEqual(rows[0]["INV_beneficiary"], "BEN")
        self.assertEqual(rows[0]["INV_region"], "EUR")
        self.assertEqual(rows[1]["INV_owner_app_name"], "NOT_GEN2")
        self.assertEqual(rows[1]["INV_beneficiary"], "NOT_GEN2")
        self.assertEqual(rows[1]["INV_region"], "NOT_GEN2")
        self.assertEqual(rows[2]["INV_owner_app_name"], "NOT_AVAILABLE")
        self.assertEqual(rows[2]["INV_beneficiary"], "NOT_AVAILABLE")
        self.assertEqual(rows[2]["INV_region"], "NOT_AVAILABLE")





    def test_build_scope_intexposed_rows_filters_all_filters_and_enriches_from_dict_dali_app(self):
        rows = [
            {
                ALL_FILTERS_FIELD: "Y",
                "server_hostname": "host-one",
                CALCULATED_SINGLE_KEAR_FIELD: "APP-ONE",
                "PCE_app": "APMA_APP-ONE_123.TRI.APP",
            },
            {
                ALL_FILTERS_FIELD: "N",
                "server_hostname": "host-two",
                CALCULATED_SINGLE_KEAR_FIELD: "APP-TWO",
                "PCE_app": "APMA_APP-TWO_123.TRI.APP",
            },
        ]
        dict_dali_app_rows = [
            {
                "uid": "APP-ONE",
                "name": "Application One",
                "short_label": "APP1",
                "dsi": "DSI",
                "application_management_rc": "RC",
            }
        ]

        scoped_rows = build_scope_intexposed_rows(rows, dict_dali_app_rows)

        self.assertEqual(len(scoped_rows), 1)
        self.assertEqual(scoped_rows[0]["server_hostname"], "host-one")
        self.assertEqual(scoped_rows[0]["name"], "Application One")
        self.assertEqual(scoped_rows[0]["short_label"], "APP1")
        self.assertEqual(scoped_rows[0]["dsi"], "DSI")
        self.assertEqual(scoped_rows[0]["application_management_rc"], "RC")
        self.assertNotIn("host-two", [row.get("server_hostname") for row in scoped_rows])


    def test_build_stats_intexposed_rows_aggregates_scope_without_enriched_block(self):
        scope_rows = [
            {
                CALCULATED_SINGLE_KEAR_FIELD: "APP-ONE",
                "exposure_scopes": "DALI.EXPOSED",
                "short_label": "APP1",
                "dsi": "DSI",
                "application_management_rc": "RC-ONE",
                "PCE_managed": "true",
                "PCE_enforcement": "full",
            },
            {
                CALCULATED_SINGLE_KEAR_FIELD: "APP-ONE",
                "exposure_scopes": "DALI.EXPOSED",
                "short_label": "APP1",
                "dsi": "DSI",
                "application_management_rc": "RC-ONE",
                "PCE_managed": "false",
                "PCE_enforcement": "idle",
            },
        ]

        stats_rows = build_stats_intexposed_rows(scope_rows)

        self.assertEqual(len(stats_rows), 1)
        self.assertEqual(stats_rows[0]["Index"], 1)
        self.assertEqual(stats_rows[0]["Program"], "INTERNET.EXPOSED (DALI.EXPOSED)")
        self.assertEqual(stats_rows[0]["Entity"], "DSI")
        self.assertEqual(stats_rows[0]["Sub-Entity"], "RC")
        self.assertEqual(stats_rows[0]["Kear ID"], "APP-ONE")
        self.assertEqual(stats_rows[0]["Application Short Label"], "APP1")
        self.assertEqual(stats_rows[0]["Total Assets in Dali (in scope)"], 2)
        self.assertEqual(stats_rows[0]["Assets in Dali not in illumio"], 1)
        self.assertEqual(stats_rows[0]["% servers with illumio installed"], "(1/2) 50,00%")
        self.assertEqual(stats_rows[0]["% servers with illumio installed Indicator Icon"], 50.0)
        self.assertEqual(stats_rows[0]["% servers with illumio installed Trend Icon"], 0.0)
        self.assertEqual(stats_rows[0]["% servers with illumio agent in blocking mode"], "(1/2) 50,00%")
        self.assertEqual(stats_rows[0]["% servers with illumio agent in blocking mode Indicator Icon"], 50.0)
        self.assertEqual(stats_rows[0]["% servers with illumio agent in blocking mode Trend Icon"], 0.0)
        for column in STATS_INTEXPOSED_COLUMNS:
            self.assertIn(column, stats_rows[0])


    def test_build_stats_intexposed_rows_keeps_multiple_kears_bucket(self):
        scope_rows = [
            {
                CALCULATED_SINGLE_KEAR_FIELD: "MULTIPLE_KEARS",
                "PCE_app": "APP-A, APP-B",
                "PCE_managed": "true",
                "PCE_enforcement": "idle",
            }
        ]

        stats_rows = build_stats_intexposed_rows(scope_rows)

        self.assertEqual(len(stats_rows), 1)
        self.assertEqual(stats_rows[0]["Program"], "INTERNET.EXPOSED")
        self.assertEqual(stats_rows[0]["Entity"], "MULTIPLE_ENTITES")
        self.assertEqual(stats_rows[0]["Sub-Entity"], "MULTIPLE_SUBENTITES")
        self.assertEqual(stats_rows[0]["Kear ID"], "MULTIPLE_KEARS")
        self.assertEqual(stats_rows[0]["Application Short Label"], "MULTIPLE_APPLICATIONS")
        self.assertEqual(stats_rows[0]["Total Assets in Dali (in scope)"], 1)
        self.assertEqual(stats_rows[0]["Assets in Dali not in illumio"], 0)
        self.assertEqual(stats_rows[0]["% servers with illumio installed"], "(1/1) 100,00%")
        self.assertEqual(stats_rows[0]["% servers with illumio agent in blocking mode"], "(0/1) 0,00%")

    def test_pce_app_label_comparison_uses_calculated_kear_dict_dali_app_pivot(self):
        rows = [
            {CALCULATED_SINGLE_KEAR_FIELD: "APP-ONE", "PCE_app": "APMA_APP-ONE_123.TRI.APP"},
            {CALCULATED_SINGLE_KEAR_FIELD: "APP-TWO", "PCE_app": "Other"},
            {CALCULATED_SINGLE_KEAR_FIELD: "APP-MISSING", "PCE_app": "APMA_APP-MISSING"},
        ]
        dict_dali_app_rows = [
            {"uid": "APP-ONE", PROPOSED_APPLICATION_LABEL_COLUMN: "APMA_APP-ONE_123.TRI.APP"},
            {"uid": "APP-TWO", PROPOSED_APPLICATION_LABEL_COLUMN: "APMA_APP-TWO_123.TRI.APP"},
        ]

        apply_pce_app_label_comparison(rows, dict_dali_app_rows)

        self.assertEqual(rows[0][PROPOSED_APPLICATION_LABEL_COLUMN], "APMA_APP-ONE_123.TRI.APP")
        self.assertEqual(rows[0]["PCE_app same as proposed"], "Y")
        self.assertEqual(rows[1][PROPOSED_APPLICATION_LABEL_COLUMN], "APMA_APP-TWO_123.TRI.APP")
        self.assertEqual(rows[1]["PCE_app same as proposed"], "N")
        self.assertEqual(rows[2][PROPOSED_APPLICATION_LABEL_COLUMN], "")
        self.assertEqual(rows[2]["PCE_app same as proposed"], "N")

    def test_pce_workload_enrichment_prefers_external_data_reference_then_hostname_fallback(self):
        rows = [
            {"server_uid": "srv-1", "server_hostname": "ignored"},
            {"server_uid": "srv-2", "server_hostname": "app01.example.net"},
            {"server_uid": "srv-3", "server_name": "10-1-2-3"},
        ]
        workload_rows = [
            {
                "external_data_reference": "SRV-1",
                "hostname": "uid-host.example.net",
                "short_hostname": "UID-HOST",
                "managed": "true",
                "enforcement": "idle",
                "app": "APP",
                "env": "PRD",
                "role": "WEB",
                "loc": "FR",
                "OS": "Linux",
                "created_at": "2026-01-01",
                "ip_with_default_gw": "192.0.2.10",
                "ocs_name_from_IP": "",
                "IPLIST": "LIST",
                "SUBNET": "192.0.2.0/24",
            },
            {"short_hostname": "APP01", "managed": "true", "hostname": "app01"},
            {"ocs_name_from_IP": "IP-10-1-2-3", "managed": "false", "hostname": "ip-host"},
        ]

        apply_pce_workload_enrichment(rows, workload_rows)

        self.assertEqual(rows[0]["PCE_match_status"], "MANAGED_WORKLOAD")
        self.assertEqual(rows[0]["PCE_match_method"], "managed external_data_reference=server_uid")
        self.assertEqual(rows[0]["PCE_hostname"], "uid-host.example.net")
        self.assertEqual(rows[0]["PCE_app"], "APP")
        self.assertEqual(rows[1]["PCE_match_status"], "MANAGED_WORKLOAD")
        self.assertEqual(rows[1]["PCE_match_method"], "managed short_hostname fallback")
        self.assertEqual(rows[2]["PCE_match_status"], "UNMANAGED_WORKLOAD")
        self.assertEqual(rows[2]["PCE_match_method"], "unmanaged ocs_name_from_IP fallback")


    def test_pce_workload_enrichment_checks_managed_rows_before_unmanaged_rows(self):
        rows = [
            {"server_uid": "srv-1", "server_hostname": "managed-host.example.net"},
            {"server_uid": "srv-2", "server_hostname": "managed-host.example.net"},
        ]
        workload_rows = [
            {
                "external_data_reference": "SRV-1",
                "short_hostname": "UNMANAGED-FIRST",
                "managed": "false",
                "hostname": "unmanaged-by-uid",
            },
            {
                "external_data_reference": "SRV-1",
                "short_hostname": "MANAGED-BY-UID",
                "managed": "true",
                "hostname": "managed-by-uid",
            },
            {"short_hostname": "MANAGED-HOST", "managed": "true", "hostname": "managed-by-hostname"},
            {"external_data_reference": "SRV-2", "managed": "false", "hostname": "unmanaged-by-uid"},
        ]

        apply_pce_workload_enrichment(rows, workload_rows)

        self.assertEqual(rows[0]["PCE_match_status"], "MANAGED_WORKLOAD")
        self.assertEqual(rows[0]["PCE_match_method"], "managed external_data_reference=server_uid")
        self.assertEqual(rows[0]["PCE_hostname"], "managed-by-uid")
        self.assertEqual(rows[1]["PCE_match_status"], "MANAGED_WORKLOAD")
        self.assertEqual(rows[1]["PCE_match_method"], "managed short_hostname fallback")
        self.assertEqual(rows[1]["PCE_hostname"], "managed-by-hostname")

    def test_platform_account_mapping_adds_ids_and_beneficiary_env(self):
        rows = [
            {"server_cloud_type": "Gen 2", "INV_owner_app_name": "ACC_A", "INV_beneficiary": "ACC_B"},
            {"server_cloud_type": "Gen 1", "INV_owner_app_name": "NOT_GEN2", "INV_beneficiary": "NOT_GEN2"},
            {"server_cloud_type": "Gen 2", "INV_owner_app_name": "NOT_AVAILABLE", "INV_beneficiary": "NOT_AVAILABLE"},
        ]
        dict_account_rows = [
            {"account": "acc_a", "id": "OWNER-1", "env": "DEV"},
            {"account": "ACC_B", "id": "BEN-1", "env": "PRD"},
        ]

        apply_platform_account_mapping(rows, dict_account_rows)

        self.assertEqual(rows[0]["PA_owner_id"], "OWNER-1")
        self.assertEqual(rows[0]["PA_beneficiary_id"], "BEN-1")
        self.assertEqual(rows[0]["PA_beneficiary_ENV"], "PRD")
        self.assertEqual(rows[1]["PA_owner_id"], "NOT_GEN2")
        self.assertEqual(rows[1]["PA_beneficiary_id"], "NOT_GEN2")
        self.assertEqual(rows[1]["PA_beneficiary_ENV"], "NOT_GEN2")
        self.assertEqual(rows[2]["PA_owner_id"], "NOT_AVAILABLE")
        self.assertEqual(rows[2]["PA_beneficiary_id"], "NOT_AVAILABLE")
        self.assertEqual(rows[2]["PA_beneficiary_ENV"], "NOT_AVAILABLE")

    def test_calculated_environment_filter_uses_platform_env_for_gen2_and_server_env_otherwise(self):
        filters = {"F_INTEXP.INCLUDE_server_environment": "PRD, UAT"}
        rows = [
            {
                "server_cloud_type": "Gen 2",
                "PA_beneficiary_ENV": "PRD",
                "server_environment": "DEV",
                "F_INTEXP.INCLUDE_server_cloud_type": "Y",
            },
            {
                "server_cloud_type": "Gen 2",
                "PA_beneficiary_ENV": "DEV",
                "server_environment": "PRD",
                "F_INTEXP.INCLUDE_server_cloud_type": "Y",
            },
            {
                "server_cloud_type": "Gen 1",
                "PA_beneficiary_ENV": "DEV",
                "server_environment": "APP-UAT-EUR",
                "F_INTEXP.INCLUDE_server_cloud_type": "Y",
            },
        ]

        apply_calculated_environment_filter(rows, filters)

        self.assertEqual(rows[0][CALCULATED_ENV_FILTER_FIELD], "Y")
        self.assertEqual(rows[0][ALL_FILTERS_FIELD], "Y")
        self.assertEqual(rows[1][CALCULATED_ENV_FILTER_FIELD], "N")
        self.assertEqual(rows[1][ALL_FILTERS_FIELD], "N")
        self.assertEqual(rows[2][CALCULATED_ENV_FILTER_FIELD], "Y")
        self.assertEqual(rows[2][ALL_FILTERS_FIELD], "Y")

    def test_distinct_inventory_accounts_uses_owner_and_beneficiary_values(self):
        rows = [
            {"INV_owner_app_name": "ACC_A", "INV_beneficiary": "ACC_B"},
            {"INV_owner_app_name": "acc_a", "INV_beneficiary": ""},
            {"INV_owner_app_name": "NOT_AVAILABLE", "INV_beneficiary": "NOT_GEN2"},
        ]

        self.assertEqual(distinct_inventory_accounts(rows), ["ACC_A", "ACC_B"])

    def test_extract_platform_tag_value_supports_id_and_env_tags(self):
        tags = ["ENV:PRD", "ID:12345"]

        self.assertEqual(extract_platform_tag_value(tags, {"ENV"}), "PRD")
        self.assertEqual(extract_platform_tag_value(tags, {"ID", "ACCOUNT_ID"}), "12345")


    def test_marley_kear_enrichment_uses_single_application_or_highest_factor(self):
        rows = [
            {"server_uid": "srv-1", "application_uid": "APP-ONE", "F_ALL_FILTERS": "Y"},
            {"server_uid": "srv-2", "application_uid": "APP-A, APP-B", "F_ALL_FILTERS": "Y"},
            {"server_uid": "srv-3", "application_uid": "APP-C, APP-D", "F_ALL_FILTERS": "Y"},
        ]
        apply_marley_kear_enrichment(
            rows,
            {
                "SRV-2": [
                    {
                        "uuid": "SRV-2",
                        "app_info": [
                            {"kear_uuid": "APP-A", "kear_factor": "40"},
                            {"kear_uuid": "APP-B", "kear_factor": "60"},
                        ],
                    }
                ],
                "SRV-3": [
                    {
                        "uuid": "SRV-3",
                        "app_info": [
                            {"kear_uuid": "APP-C", "kear_factor": "50"},
                            {"kear_uuid": "APP-D", "kear_factor": "50"},
                        ],
                    }
                ],
            },
        )

        self.assertEqual(rows[0][CALCULATED_SINGLE_KEAR_FIELD], "APP-ONE")
        self.assertEqual(rows[1][CALCULATED_SINGLE_KEAR_FIELD], "APP-B")
        self.assertEqual(rows[1]["MAR_app_info.kear_uuid"], "APP-A, APP-B")
        self.assertEqual(rows[1]["MAR_app_info.kear_factor"], "40, 60")
        self.assertEqual(rows[2][CALCULATED_SINGLE_KEAR_FIELD], "MULTIPLE_KEARS")

    def test_marley_kear_enrichment_keeps_duplicate_factor_values(self):
        rows = [{"server_uid": "srv-4", "application_uid": "APP-A, APP-B, APP-C", "F_ALL_FILTERS": "Y"}]

        apply_marley_kear_enrichment(
            rows,
            {
                "SRV-4": [
                    {
                        "uuid": "SRV-4",
                        "app_info": {
                            "kear_uuid": ["APP-A", "APP-B", "APP-C"],
                            "kear_factor": [34, 33, 33],
                        },
                    }
                ]
            },
        )

        self.assertEqual(rows[0]["MAR_app_info.kear_uuid"], "APP-A, APP-B, APP-C")
        self.assertEqual(rows[0]["MAR_app_info.kear_factor"], "34, 33, 33")
        self.assertEqual(rows[0][CALCULATED_SINGLE_KEAR_FIELD], "APP-A")

    def test_marley_kear_enrichment_recovers_empty_application_uid_and_marks_missing(self):
        rows = [
            {"server_uid": "srv-5", "application_uid": "", "F_ALL_FILTERS": "Y"},
            {"server_uid": "srv-6", "application_uid": "APP-X, APP-Y", "F_ALL_FILTERS": "Y"},
        ]

        apply_marley_kear_enrichment(
            rows,
            {
                "SRV-5": [
                    {
                        "uuid": "SRV-5",
                        "app_info": {
                            "kear_uuid": ["APP-Z"],
                            "kear_factor": [100],
                        },
                    }
                ]
            },
        )

        self.assertEqual(rows[0]["MAR_app_info.kear_uuid"], "APP-Z")
        self.assertEqual(rows[0]["MAR_app_info.kear_factor"], "100")
        self.assertEqual(rows[0][CALCULATED_SINGLE_KEAR_FIELD], "APP-Z")
        self.assertEqual(rows[1][CALCULATED_SINGLE_KEAR_FIELD], MISSING_KEAR_VALUE)

    def test_collect_dict_dali_app_uids_keeps_sources_for_filtered_rows_only(self):
        rows = [
            {ALL_FILTERS_FIELD: "Y", CALCULATED_SINGLE_KEAR_FIELD: "APP-ONE", "MAR_app_info.kear_uuid": ""},
            {
                ALL_FILTERS_FIELD: "Y",
                CALCULATED_SINGLE_KEAR_FIELD: "MULTIPLE_KEARS",
                "MAR_app_info.kear_uuid": "APP-TWO, APP-THREE",
            },
            {ALL_FILTERS_FIELD: "N", CALCULATED_SINGLE_KEAR_FIELD: "APP-IGNORED", "MAR_app_info.kear_uuid": ""},
            {ALL_FILTERS_FIELD: "Y", CALCULATED_SINGLE_KEAR_FIELD: MISSING_KEAR_VALUE, "MAR_app_info.kear_uuid": ""},
        ]

        uid_rows = collect_dict_dali_app_uids(rows)

        self.assertEqual(
            uid_rows,
            [
                {"uid": "APP-ONE", DICT_DALI_APP_SOURCE_FIELD: CALCULATED_SINGLE_KEAR_FIELD},
                {"uid": "APP-TWO", DICT_DALI_APP_SOURCE_FIELD: "MULTIPLE_KEARS"},
                {"uid": "APP-THREE", DICT_DALI_APP_SOURCE_FIELD: "MULTIPLE_KEARS"},
            ],
        )

    def test_build_dict_dali_app_rows_uses_dali_search_properties(self):
        class FakeClient:
            def __init__(self):
                self.payloads = []

            def post_json(self, endpoint, payload):
                self.payloads.append((endpoint, payload))
                uid = payload["filters"][0]["attributeValue"]
                return {
                    "count": 1,
                    "result": [
                        {
                            "leading_node": {
                                "properties": {
                                    "uid": uid,
                                    "name": f"Application {uid}",
                                    "status": "In use",
                                }
                            }
                        }
                    ],
                }

        rows = build_dict_dali_app_rows(
            FakeClient(),
            [{"uid": "APP-ONE", DICT_DALI_APP_SOURCE_FIELD: CALCULATED_SINGLE_KEAR_FIELD}],
            "/api/v1/search",
        )

        self.assertEqual([row["uid"] for row in rows], ["APP-ONE"])
        self.assertEqual(rows[0][DICT_DALI_APP_SOURCE_FIELD], CALCULATED_SINGLE_KEAR_FIELD)
        self.assertEqual(rows[0]["name"], "Application APP-ONE")
        self.assertEqual(rows[0]["status"], "In use")
        self.assertEqual(set(APPLICATION_DICTIONARY_HEADERS), set(rows[0].keys()))

    def test_extract_identifier_pairs_supports_nested_and_dotted_kear_appli_docs(self):
        issuers, identifiers = extract_identifier_pairs(
            {"identifiers": [{"issuer": "IRT", "identifier": "123"}, {"issuer": "IAPPLI", "identifier": "APP"}]}
        )

        self.assertEqual(issuers, ["IRT", "IAPPLI"])
        self.assertEqual(identifiers, ["123", "APP"])

        dotted_issuers, dotted_identifiers = extract_identifier_pairs(
            {"identifiers.issuer": ["IRT", "IAPPLI (Trigram)"], "identifiers.identifier": ["123", "TRI"]}
        )

        self.assertEqual(dotted_issuers, ["IRT", "IAPPLI (Trigram)"])
        self.assertEqual(dotted_identifiers, ["123", "TRI"])

    def test_build_proposed_application_label_keeps_w05_order(self):
        label = build_proposed_application_label(
            "APP-ONE",
            ["IAPPLI", "IRT", "IAPPLI (Trigram)", "IGNORED"],
            ["APP", "123", "TRI", "NOPE"],
        )

        self.assertEqual(label, "APMA_APP-ONE_123.TRI.APP")

    def test_enrich_dict_dali_app_rows_with_kear_appli_adds_w05_columns(self):
        rows = [{"uid": "APP-ONE", DICT_DALI_APP_SOURCE_FIELD: CALCULATED_SINGLE_KEAR_FIELD}]

        with patch(
            "internet_exposed_extract.query_kear_appli_by_global_ids",
            return_value={
                "APP-ONE": {
                    "global_id": "APP-ONE",
                    "identifiers": [
                        {"issuer": "IRT", "identifier": "123"},
                        {"issuer": "IAPPLI (Trigram)", "identifier": "TRI"},
                        {"issuer": "IAPPLI", "identifier": "APP"},
                    ],
                }
            },
        ):
            enrich_dict_dali_app_rows_with_kear_appli(rows)

        self.assertEqual(rows[0][KEAR_APPLI_ISSUER_COLUMN], "IRT, IAPPLI (Trigram), IAPPLI")
        self.assertEqual(rows[0][KEAR_APPLI_IDENTIFIER_COLUMN], "123, TRI, APP")
        self.assertEqual(rows[0][PROPOSED_APPLICATION_LABEL_COLUMN], "APMA_APP-ONE_123.TRI.APP")

    def test_query_kear_appli_by_global_ids_uses_normalized_lookup_values(self):
        class FakeData4SecClient:
            def __init__(self):
                self.es_connection = object()

            def bulk_search_multi(self, **kwargs):
                self.kwargs = kwargs
                self.__class__.last_kwargs = kwargs
                return {
                    "APP-ONE": [
                        {
                            "global_id": "APP-ONE",
                            "identifiers": [{"issuer": "IRT", "identifier": "123"}],
                        }
                    ]
                }

        with patch("internet_exposed_extract.Data4secClient", FakeData4SecClient):
            docs_by_uid = query_kear_appli_by_global_ids(["app-one"])

        self.assertEqual(FakeData4SecClient.last_kwargs["values"], ["APP-ONE"])
        self.assertIn("APP-ONE", docs_by_uid)
        self.assertEqual(docs_by_uid["APP-ONE"]["global_id"], "APP-ONE")

    def test_dict_account_sheet_has_formatting(self):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed in this environment")

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "internet_exposed.xlsx"
            write_xlsx(
                path,
                rows=[
                    {
                        "server_uid": "srv-1",
                        "PCE_app": "APMA_APP-ONE_123.TRI.APP",
                        "INV_owner_app_name": "Owner",
                        "PA_owner_id": "OWNER-ID",
                        PROPOSED_APPLICATION_LABEL_COLUMN: "APMA_APP-ONE_123.TRI.APP",
                        "PCE_app same as proposed": "Y",
                        "F_ALL_FILTERS": "Y",
                        CALCULATED_SINGLE_KEAR_FIELD: "APP-ONE",
                        "PCE_managed": "true",
                        "PCE_enforcement": "full",
                    },
                    {"server_uid": "srv-2", "F_ALL_FILTERS": "N", CALCULATED_SINGLE_KEAR_FIELD: "APP-TWO"},
                ],
                fieldnames=[
                    "server_uid",
                    "INV_owner_app_name",
                    "PA_owner_id",
                    "PCE_app",
                    PROPOSED_APPLICATION_LABEL_COLUMN,
                    "PCE_app same as proposed",
                    "F_ALL_FILTERS",
                    "PCE_app",
                ],
                dict_account_rows=[{"account": "ACC_A", "id": "123", "env": "PRD"}],
                dict_dali_app_rows=[
                    {"uid": "APP-ONE", DICT_DALI_APP_SOURCE_FIELD: CALCULATED_SINGLE_KEAR_FIELD, "name": "Application One"}
                ],
            )
            workbook = load_workbook(path)
            try:
                raw_ws = workbook["RAW_INTERNET_EXPOSED"]
                self.assertEqual(raw_ws.max_column, 7)
                self.assertEqual(raw_ws["B1"].fill.fgColor.rgb, "005B9BD5")
                self.assertEqual(raw_ws["C1"].fill.fgColor.rgb, "005B9BD5")
                self.assertEqual(raw_ws["D1"].fill.fgColor.rgb, "00F4B183")
                self.assertEqual(raw_ws["E1"].value, PROPOSED_APPLICATION_LABEL_COLUMN)
                self.assertEqual(raw_ws["F1"].value, "PCE_app same as proposed")
                self.assertEqual(raw_ws["E1"].fill.fgColor.rgb, "0070AD47")
                self.assertEqual(raw_ws["F1"].fill.fgColor.rgb, "0070AD47")
                self.assertNotIn("STATS", workbook.sheetnames)
                stats_ws = workbook[STATS_INTEXPOSED_SHEET]
                self.assertEqual([cell.value for cell in stats_ws[1]], STATS_INTEXPOSED_COLUMNS)
                self.assertEqual(stats_ws.cell(row=2, column=STATS_INTEXPOSED_COLUMNS.index("Kear ID") + 1).value, "APP-ONE")
                self.assertEqual(
                    stats_ws.cell(row=2, column=STATS_INTEXPOSED_COLUMNS.index("% servers with illumio installed") + 1).value,
                    "(1/1) 100,00%",
                )
                self.assertEqual(
                    stats_ws.cell(row=2, column=STATS_INTEXPOSED_COLUMNS.index("% servers with illumio installed Indicator Icon") + 1).value,
                    100,
                )
                scope_ws = workbook[SCOPE_INTEXPOSED_SHEET]
                self.assertEqual([cell.value for cell in scope_ws[1]], SCOPE_INTEXPOSED_FIELDS)
                self.assertEqual(scope_ws.max_row, 2)
                self.assertEqual(scope_ws.cell(row=2, column=SCOPE_INTEXPOSED_FIELDS.index("server_uid") + 1).value, "srv-1")
                self.assertEqual(scope_ws.cell(row=2, column=SCOPE_INTEXPOSED_FIELDS.index("name") + 1).value, "Application One")
                self.assertEqual(
                    scope_ws.cell(row=1, column=SCOPE_INTEXPOSED_FIELDS.index("name") + 1).fill.fgColor.rgb,
                    "00C65911",
                )
                self.assertEqual(
                    scope_ws.cell(row=1, column=SCOPE_INTEXPOSED_FIELDS.index("INV_owner_app_name") + 1).fill.fgColor.rgb,
                    "005B9BD5",
                )
                self.assertEqual(set(INV_PA_HIGHLIGHT_FIELDS).issubset(set(SCOPE_INTEXPOSED_FIELDS)), True)
                ws = workbook["DictAccount"]
                self.assertEqual([cell.value for cell in ws[1]], ["account", "id", "env"])
                self.assertEqual(ws.freeze_panes, "A2")
                self.assertEqual(ws.auto_filter.ref, ws.dimensions)
                self.assertEqual(ws["A1"].fill.fgColor.rgb, "008064A2")
                self.assertIsNotNone(ws["A2"].border.left.style)
                self.assertGreaterEqual(ws.column_dimensions["A"].width, 14)
                app_ws = workbook["DictDaliApp"]
                self.assertEqual([cell.value for cell in app_ws[1]], APPLICATION_DICTIONARY_HEADERS)
                self.assertEqual(app_ws["A2"].value, "APP-ONE")
                self.assertEqual(app_ws["B2"].value, CALCULATED_SINGLE_KEAR_FIELD)
                self.assertEqual(app_ws.freeze_panes, "A2")
            finally:
                workbook.close()

    def test_read_filters_conf_supports_equal_and_comma_separators(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "filters.conf"
            path.write_text("F_A=one,two\nF_B,three,four\n", encoding="utf-8")
            filters = read_filters_conf(str(path))

        self.assertEqual(filters["F_A"], "one,two")
        self.assertEqual(filters["F_B"], "three,four")


if __name__ == "__main__":
    unittest.main()
