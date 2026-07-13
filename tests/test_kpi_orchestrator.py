import logging
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from kpi_orchestrator import (
    SCOPE_INTEXPOSED_SHEET,
    append_internet_exposed_stats_to_kpi_workbook,
    append_internet_exposed_totals_to_kpi_workbook,
    maybe_send_kpi_email,
    rotate_root_logs_once_per_week,
    _infer_stats_headers,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill
except ImportError:  # pragma: no cover - optional test dependency
    Workbook = None
    load_workbook = None
    PatternFill = None


class InferStatsHeadersTests(unittest.TestCase):
    def test_scope_intexposed_sheet_constant_matches_generated_workbook_sheet(self):
        self.assertEqual(SCOPE_INTEXPOSED_SHEET, "SCOPE.INTEXPOSED")

    def test_hidden_icon_headers_are_inferred_from_percentage_columns(self):
        headers = _infer_stats_headers(
            [
                "Index",
                "% servers with illumio installed",
                None,
                "",
                "% servers with illumio installed (Enriched)",
                None,
                "",
            ]
        )

        self.assertEqual(
            headers,
            [
                "Index",
                "% servers with illumio installed",
                "% servers with illumio installed Indicator Icon",
                "% servers with illumio installed Trend Icon",
                "% servers with illumio installed (Enriched)",
                "% servers with illumio installed (Enriched) Indicator Icon",
                "% servers with illumio installed (Enriched) Trend Icon",
            ],
        )


@unittest.skipIf(Workbook is None or load_workbook is None or PatternFill is None, "openpyxl is required for XLSX append tests")
class AppendInternetExposedStatsTests(unittest.TestCase):
    def test_append_maps_stats_columns_and_sets_enriched_values_to_na(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            kpi_path = tmp_path / "kpi_microseg.xlsx"
            internet_path = tmp_path / "internet_exposed.xlsx"

            kpi_wb = Workbook()
            stats_ws = kpi_wb.active
            stats_ws.title = "STATS"
            stats_ws.append(
                [
                    "Index",
                    "Program",
                    "Kear ID",
                    "Total Assets in Dali (in scope)",
                    "Total Assets in Dali (Enriched)",
                    "% servers with illumio installed",
                    "% servers with illumio installed (Enriched)",
                    "% servers with illumio installed Indicator Icon",
                    "% servers with illumio installed (Enriched) Indicator Icon",
                ]
            )
            stats_ws.append([1, "P1", "APP-BASE", "2", "3", "(1/2) 50,00%", "(2/3) 66,67%", 50, 66.67])
            stats_ws.auto_filter.ref = stats_ws.dimensions
            kpi_wb.save(kpi_path)
            kpi_wb.close()

            internet_wb = Workbook()
            internet_ws = internet_wb.active
            internet_ws.title = "STATS.INTEXPOSED"
            internet_ws.append(
                [
                    "Index",
                    "Program",
                    "Kear ID",
                    "Total Assets in Dali (in scope)",
                    "% servers with illumio installed",
                    "% servers with illumio installed Indicator Icon",
                ]
            )
            internet_ws.append([1, "PINT", "app-int", "1", "(1/1) 100,00%", 100])
            internet_ws.append([2, "PINT", "APP-INT", "1", "(1/1) 100,00%", 100])
            internet_wb.save(internet_path)
            internet_wb.close()

            appended = append_internet_exposed_stats_to_kpi_workbook(
                kpi_xlsx=kpi_path,
                internet_exposed_xlsx=internet_path,
                log=logging.getLogger("test"),
            )

            self.assertEqual(appended, 1)
            appended_again = append_internet_exposed_stats_to_kpi_workbook(
                kpi_xlsx=kpi_path,
                internet_exposed_xlsx=internet_path,
                log=logging.getLogger("test"),
            )
            self.assertEqual(appended_again, 0)
            workbook = load_workbook(kpi_path)
            try:
                stats_ws = workbook["STATS"]
                self.assertEqual(stats_ws.max_row, 3)
                self.assertEqual(stats_ws.cell(row=3, column=1).value, 2)
                self.assertEqual(stats_ws.cell(row=3, column=2).value, "PINT")
                self.assertEqual(stats_ws.cell(row=3, column=3).value, "APP-INT")
                self.assertEqual(stats_ws.cell(row=3, column=4).value, "1")
                self.assertEqual(stats_ws.cell(row=3, column=5).value, "N/A")
                self.assertEqual(stats_ws.cell(row=3, column=6).value, "(1/1) 100,00%")
                self.assertEqual(stats_ws.cell(row=3, column=7).value, "N/A")
                self.assertEqual(stats_ws.cell(row=3, column=8).value, 100)
                self.assertEqual(stats_ws.cell(row=3, column=9).value, "N/A")
                self.assertEqual(stats_ws.auto_filter.ref, "A1:I3")
            finally:
                workbook.close()

    def test_mail_attachment_workbook_receives_internet_exposed_stats_append(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            raw_dir = tmp_path / "raw"
            raw_dir.mkdir()
            output_path = raw_dir / "dali_impact_analysis_20260101_000000.xlsx"
            internet_path = raw_dir / "internet_exposed_20260101_000000.xlsx"
            pptx_path = raw_dir / "dali_impact_analysis_20260101_000000.pptx"
            pptx_path.write_bytes(b"pptx")

            source_wb = Workbook()
            stats_ws = source_wb.active
            stats_ws.title = "STATS"
            stats_ws.append(["Index", "Program", "Kear ID", "Total Assets in Dali (Enriched)"])
            stats_ws.append([1, "P1", "APP-BASE", "3"])
            for sheet_name in [
                "Summary",
                "SCOPE",
                "TOTAL.PROGRAM",
                "TOTAL.ENTITY",
                "NOT_IN_ILLUMIO",
                "IN_ILLUMIO_BUT_NOT_BLOCKING",
                "EXCLUDED",
                "GLOBAL",
                "OUT_OF_SCOPE",
                "MONITORED_SCOPES",
            ]:
                ws = source_wb.create_sheet(sheet_name)
                ws.append(["Header"])
            source_wb.save(output_path)
            source_wb.close()

            internet_wb = Workbook()
            internet_ws = internet_wb.active
            internet_ws.title = "STATS.INTEXPOSED"
            internet_ws.append(["Index", "Program", "Kear ID"])
            internet_ws.append([1, "PINT", "app-int"])
            scope_ws = internet_wb.create_sheet("SCOPE.INTEXPOSED")
            scope_ws.append(["scope_col"])
            scope_ws.append(["scope_value"])
            scope_ws["A1"].fill = PatternFill("solid", fgColor="5B9BD5")
            scope_ws["A2"].fill = PatternFill("solid", fgColor="DDEBF7")
            internet_wb.save(internet_path)
            internet_wb.close()

            captured_attachments = []

            def fake_send_carto_notification(**kwargs):
                captured_attachments.extend(kwargs["attachment_paths"])

            env = {
                "MAIL_TO": "recipient@example.com",
                "SMTP_SERVER": "smtp.example.com",
            }
            with patch.dict(os.environ, env, clear=False), patch(
                "kpi_orchestrator.send_carto_notification",
                side_effect=fake_send_carto_notification,
            ):
                maybe_send_kpi_email(
                    timestamp="20260101_000000",
                    raw_dir=raw_dir,
                    output_xlsx=output_path,
                    internet_exposed_xlsx=internet_path,
                    meta={},
                    log=logging.getLogger("test"),
                )

            slim_path = raw_dir / "kpi_microseg_20260101_000000.xlsx"
            self.assertTrue(slim_path.is_file())
            self.assertIn(slim_path, captured_attachments)
            workbook = load_workbook(slim_path)
            try:
                sheet_names = workbook.sheetnames
                self.assertEqual(sheet_names[sheet_names.index("SCOPE") + 1], "SCOPE.INTEXPOSED")
                scope_ws = workbook["SCOPE.INTEXPOSED"]
                self.assertEqual(scope_ws["A1"].value, "scope_col")
                self.assertEqual(scope_ws["A2"].value, "scope_value")
                self.assertEqual(scope_ws["A1"].fill.fgColor.rgb, "005B9BD5")
                self.assertEqual(scope_ws["A2"].fill.fgColor.rgb, "00DDEBF7")
                stats_ws = workbook["STATS"]
                self.assertEqual(stats_ws.max_row, 3)
                self.assertEqual(stats_ws.cell(row=3, column=1).value, 2)
                self.assertEqual(stats_ws.cell(row=3, column=2).value, "PINT")
                self.assertEqual(stats_ws.cell(row=3, column=3).value, "APP-INT")
                self.assertEqual(stats_ws.cell(row=3, column=4).value, "N/A")
            finally:
                workbook.close()

    def test_append_enriches_total_program_and_total_entity_with_single_internet_exposed_program(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            kpi_path = tmp_path / "kpi.xlsx"
            internet_path = tmp_path / "internet.xlsx"

            kpi_wb = Workbook()
            total_program = kpi_wb.active
            total_program.title = "TOTAL.PROGRAM"
            total_program.append(
                [
                    "Program",
                    "Number of Applications",
                    "Total Assets in Dali (in scope)",
                    "% servers with illumio installed",
                    "% servers with illumio agent in blocking mode",
                ]
            )
            total_program.append(["BASE", "1", "2", "(1/2) 50,00%", "(0/2) 0,00%"])
            total_entity = kpi_wb.create_sheet("TOTAL.ENTITY")
            total_entity.append(
                [
                    "Entity",
                    "Program",
                    "Number of Applications",
                    "Total Assets in Dali (in scope)",
                    "% servers with illumio installed",
                    "% servers with illumio agent in blocking mode",
                ]
            )
            kpi_wb.save(kpi_path)
            kpi_wb.close()

            internet_wb = Workbook()
            stats = internet_wb.active
            stats.title = "STATS.INTEXPOSED"
            stats.append(
                [
                    "Index",
                    "Program",
                    "Entity",
                    "Kear ID",
                    "Total Assets in Dali (in scope)",
                    "% servers with illumio installed",
                    "% servers with illumio agent in blocking mode",
                ]
            )
            stats.append([1, "INTERNET.EXPOSED (DALI.EXPOSED)", "DSI-A", "app-one", 2, "(1/2) 50,00%", "(1/2) 50,00%"])
            stats.append([2, "INTERNET.EXPOSED (MASAI.EXPOSED)", "DSI-A", "app-two", 3, "(3/3) 100,00%", "(0/3) 0,00%"])
            internet_wb.save(internet_path)
            internet_wb.close()

            appended = append_internet_exposed_totals_to_kpi_workbook(
                kpi_xlsx=kpi_path, internet_exposed_xlsx=internet_path, log=logging.getLogger("test")
            )
            self.assertEqual(appended, 2)
            self.assertEqual(
                append_internet_exposed_totals_to_kpi_workbook(
                    kpi_xlsx=kpi_path, internet_exposed_xlsx=internet_path, log=logging.getLogger("test")
                ),
                0,
            )

            workbook = load_workbook(kpi_path)
            try:
                total_program = workbook["TOTAL.PROGRAM"]
                self.assertEqual(total_program.cell(row=3, column=1).value, "INTERNET.EXPOSED")
                self.assertEqual(total_program.cell(row=3, column=2).value, "2")
                self.assertEqual(total_program.cell(row=3, column=3).value, "5")
                self.assertEqual(total_program.cell(row=3, column=4).value, "(4/5) 80,00%")
                self.assertEqual(total_program.cell(row=3, column=5).value, "(1/5) 20,00%")
                total_entity = workbook["TOTAL.ENTITY"]
                self.assertEqual(total_entity.cell(row=2, column=1).value, "DSI-A")
                self.assertEqual(total_entity.cell(row=2, column=2).value, "INTERNET.EXPOSED")
            finally:
                workbook.close()

    def test_rotate_root_logs_compresses_week_old_root_logs_into_logs_folder(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            source = tmp_path / "workloader.log"
            source.write_text("old log", encoding="utf-8")
            old_time = 1_700_000_000
            os.utime(source, (old_time, old_time))

            archived = rotate_root_logs_once_per_week(tmp_path, logging.getLogger("test"))

            self.assertEqual(len(archived), 1)
            self.assertEqual(archived[0].parent, tmp_path / "logs")
            self.assertTrue(archived[0].name.startswith("workloader."))
            self.assertEqual(source.read_text(encoding="utf-8"), "")


if __name__ == "__main__":
    unittest.main()
